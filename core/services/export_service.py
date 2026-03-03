# core/services/export_service.py
"""
Service d'export pour la génération de fichiers PDF, XML, CSV, Excel.
Utilise StreamingHttpResponse pour les fichiers volumineux.
"""
import io
import csv
import json
from decimal import Decimal
from datetime import datetime, date
from typing import Generator, Any, Dict, List, Optional, Union
from xml.etree.ElementTree import Element, SubElement, tostring
from xml.dom import minidom

from django.http import StreamingHttpResponse, HttpResponse, FileResponse
from django.core.files.base import ContentFile
from django.utils.translation import gettext as _

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


class ExportService:
    """
    Service centralisé pour tous les exports de fichiers.
    """

    # Constantes
    CHUNK_SIZE = 8192  # Taille des chunks pour le streaming

    # ========================================================================
    # MÉTHODES GÉNÉRIQUES
    # ========================================================================

    @staticmethod
    def get_content_type(format_type: str) -> str:
        """Retourne le content-type approprié pour un format."""
        content_types = {
            'pdf': 'application/pdf',
            'csv': 'text/csv; charset=utf-8',
            'excel': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xml': 'application/xml',
            'json': 'application/json',
        }
        return content_types.get(format_type.lower(), 'application/octet-stream')

    @staticmethod
    def get_file_extension(format_type: str) -> str:
        """Retourne l'extension de fichier appropriée."""
        extensions = {
            'pdf': 'pdf',
            'csv': 'csv',
            'excel': 'xlsx',
            'xlsx': 'xlsx',
            'xml': 'xml',
            'json': 'json',
        }
        return extensions.get(format_type.lower(), 'bin')

    @staticmethod
    def serialize_value(value: Any) -> str:
        """Sérialise une valeur pour l'export."""
        if value is None:
            return ''
        if isinstance(value, Decimal):
            return str(value)
        if isinstance(value, (datetime, date)):
            return value.isoformat() if isinstance(value, datetime) else value.strftime('%Y-%m-%d')
        if isinstance(value, bool):
            return 'Oui' if value else 'Non'
        return str(value)

    # ========================================================================
    # CSV EXPORT (avec StreamingHttpResponse)
    # ========================================================================

    @classmethod
    def generate_csv_response(
        cls,
        data: List[Dict],
        headers: List[str],
        filename: str,
        delimiter: str = ';'
    ) -> StreamingHttpResponse:
        """
        Génère une réponse CSV en streaming.

        Args:
            data: Liste de dictionnaires avec les données
            headers: Liste des en-têtes de colonnes
            filename: Nom du fichier
            delimiter: Délimiteur CSV (défaut: ';' pour compatibilité Excel)

        Returns:
            StreamingHttpResponse avec le fichier CSV
        """
        def csv_generator() -> Generator[str, None, None]:
            # BOM UTF-8 pour Excel
            yield '\ufeff'

            # En-têtes
            output = io.StringIO()
            writer = csv.writer(output, delimiter=delimiter, quoting=csv.QUOTE_MINIMAL)
            writer.writerow(headers)
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

            # Données
            for row in data:
                values = [cls.serialize_value(row.get(h, '')) for h in headers]
                writer.writerow(values)
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)

        response = StreamingHttpResponse(
            csv_generator(),
            content_type=cls.get_content_type('csv')
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response

    @classmethod
    def generate_csv_from_queryset(
        cls,
        queryset,
        fields: List[str],
        field_labels: Optional[Dict[str, str]] = None,
        filename: str = 'export'
    ) -> StreamingHttpResponse:
        """
        Génère un CSV en streaming depuis un QuerySet Django.

        Args:
            queryset: QuerySet Django
            fields: Liste des champs à exporter
            field_labels: Mapping champ -> label (optionnel)
            filename: Nom du fichier

        Returns:
            StreamingHttpResponse avec le fichier CSV
        """
        field_labels = field_labels or {}

        def csv_generator() -> Generator[str, None, None]:
            yield '\ufeff'

            output = io.StringIO()
            writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)

            # En-têtes
            headers = [field_labels.get(f, f) for f in fields]
            writer.writerow(headers)
            yield output.getvalue()
            output.seek(0)
            output.truncate(0)

            # Données en chunks pour économiser la mémoire
            for obj in queryset.iterator(chunk_size=500):
                values = []
                for field in fields:
                    # Support des relations (ex: 'client__nom')
                    value = obj
                    for part in field.split('__'):
                        value = getattr(value, part, None)
                        if value is None:
                            break
                    values.append(cls.serialize_value(value))

                writer.writerow(values)
                yield output.getvalue()
                output.seek(0)
                output.truncate(0)

        response = StreamingHttpResponse(
            csv_generator(),
            content_type=cls.get_content_type('csv')
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response

    # ========================================================================
    # EXCEL EXPORT (avec StreamingHttpResponse)
    # ========================================================================

    @classmethod
    def generate_excel_response(
        cls,
        data: List[Dict],
        headers: List[str],
        filename: str,
        sheet_name: str = 'Export'
    ) -> HttpResponse:
        """
        Génère une réponse Excel.

        Args:
            data: Liste de dictionnaires avec les données
            headers: Liste des en-têtes de colonnes
            filename: Nom du fichier
            sheet_name: Nom de la feuille

        Returns:
            HttpResponse avec le fichier Excel
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-têtes
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Données
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, '')
                if isinstance(value, Decimal):
                    value = float(value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border

        # Auto-ajuster les colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type=cls.get_content_type('excel')
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    @classmethod
    def generate_excel_streaming(
        cls,
        queryset,
        fields: List[str],
        field_labels: Optional[Dict[str, str]] = None,
        filename: str = 'export',
        sheet_name: str = 'Export'
    ) -> StreamingHttpResponse:
        """
        Génère un Excel en streaming pour les gros volumes.

        Note: Pour de très gros fichiers, utiliser xlsxwriter avec streaming.
        """
        field_labels = field_labels or {}

        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_name)

        # En-têtes
        headers = [field_labels.get(f, f) for f in fields]
        ws.append(headers)

        # Données
        for obj in queryset.iterator(chunk_size=500):
            values = []
            for field in fields:
                value = obj
                for part in field.split('__'):
                    value = getattr(value, part, None)
                    if value is None:
                        break
                if isinstance(value, Decimal):
                    value = float(value)
                values.append(cls.serialize_value(value) if value else '')
            ws.append(values)

        # Sauvegarder
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        def file_iterator():
            while chunk := buffer.read(cls.CHUNK_SIZE):
                yield chunk

        response = StreamingHttpResponse(
            file_iterator(),
            content_type=cls.get_content_type('excel')
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    # ========================================================================
    # XML EXPORT (avec StreamingHttpResponse)
    # ========================================================================

    @classmethod
    def generate_xml_response(
        cls,
        data: Union[Dict, List],
        root_name: str = 'data',
        item_name: str = 'item',
        filename: str = 'export',
        pretty: bool = True
    ) -> StreamingHttpResponse:
        """
        Génère une réponse XML en streaming.

        Args:
            data: Données à exporter (dict ou list)
            root_name: Nom de l'élément racine
            item_name: Nom des éléments pour les listes
            filename: Nom du fichier
            pretty: Formater le XML pour lisibilité

        Returns:
            StreamingHttpResponse avec le fichier XML
        """
        def build_xml_element(parent: Element, key: str, value: Any):
            """Construit récursivement les éléments XML."""
            if isinstance(value, dict):
                child = SubElement(parent, key)
                for k, v in value.items():
                    build_xml_element(child, k, v)
            elif isinstance(value, list):
                child = SubElement(parent, key)
                for item in value:
                    if isinstance(item, dict):
                        item_elem = SubElement(child, item_name)
                        for k, v in item.items():
                            build_xml_element(item_elem, k, v)
                    else:
                        SubElement(child, item_name).text = cls.serialize_value(item)
            else:
                SubElement(parent, key).text = cls.serialize_value(value)

        # Construire le XML
        root = Element(root_name)

        if isinstance(data, dict):
            for key, value in data.items():
                build_xml_element(root, key, value)
        elif isinstance(data, list):
            for item in data:
                if isinstance(item, dict):
                    item_elem = SubElement(root, item_name)
                    for k, v in item.items():
                        build_xml_element(item_elem, k, v)
                else:
                    SubElement(root, item_name).text = cls.serialize_value(item)

        # Convertir en string
        xml_bytes = tostring(root, encoding='utf-8')

        if pretty:
            reparsed = minidom.parseString(xml_bytes)
            xml_bytes = reparsed.toprettyxml(indent="  ", encoding='utf-8')

        def xml_generator():
            yield xml_bytes

        response = StreamingHttpResponse(
            xml_generator(),
            content_type=cls.get_content_type('xml')
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xml"'
        return response

    # ========================================================================
    # JSON EXPORT (avec StreamingHttpResponse)
    # ========================================================================

    @classmethod
    def generate_json_response(
        cls,
        data: Any,
        filename: str = 'export',
        indent: int = 2
    ) -> StreamingHttpResponse:
        """
        Génère une réponse JSON en streaming.

        Args:
            data: Données à exporter
            filename: Nom du fichier
            indent: Indentation pour le formatage

        Returns:
            StreamingHttpResponse avec le fichier JSON
        """
        def json_encoder(obj):
            """Encoder personnalisé pour les types non-JSON."""
            if isinstance(obj, Decimal):
                return float(obj)
            if isinstance(obj, (datetime, date)):
                return obj.isoformat()
            raise TypeError(f"Object of type {type(obj)} is not JSON serializable")

        json_str = json.dumps(data, ensure_ascii=False, indent=indent, default=json_encoder)

        def json_generator():
            yield json_str.encode('utf-8')

        response = StreamingHttpResponse(
            json_generator(),
            content_type=cls.get_content_type('json')
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.json"'
        return response

    # ========================================================================
    # PDF EXPORT
    # ========================================================================

    @classmethod
    def generate_pdf_response(
        cls,
        content_builder: callable,
        filename: str = 'export'
    ) -> HttpResponse:
        """
        Génère une réponse PDF.

        Args:
            content_builder: Fonction qui prend un canvas et y dessine le contenu
            filename: Nom du fichier

        Returns:
            HttpResponse avec le fichier PDF
        """
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)

        # Appeler le builder pour dessiner le contenu
        content_builder(p)

        p.save()
        buffer.seek(0)

        response = HttpResponse(buffer.getvalue(), content_type=cls.get_content_type('pdf'))
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response

    @classmethod
    def generate_pdf_streaming(
        cls,
        content_builder: callable,
        filename: str = 'export'
    ) -> StreamingHttpResponse:
        """
        Génère une réponse PDF en streaming pour les gros fichiers.
        """
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer, pagesize=A4)

        content_builder(p)

        p.save()
        buffer.seek(0)

        def pdf_generator():
            while chunk := buffer.read(cls.CHUNK_SIZE):
                yield chunk

        response = StreamingHttpResponse(
            pdf_generator(),
            content_type=cls.get_content_type('pdf')
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response

    # ========================================================================
    # EXPORT COMPTABLE (Balance, Grand Livre, etc.)
    # ========================================================================

    @classmethod
    def export_balance(
        cls,
        comptes: List[Dict],
        titre: str,
        date_debut: date,
        date_fin: date,
        format_type: str = 'pdf'
    ) -> Union[StreamingHttpResponse, HttpResponse]:
        """
        Exporte une balance au format spécifié.

        Args:
            comptes: Liste des comptes avec soldes
            titre: Titre du document
            date_debut, date_fin: Période
            format_type: 'pdf', 'csv', 'excel'
        """
        headers = ['Numéro', 'Libellé', 'Débit', 'Crédit', 'Solde']

        if format_type == 'csv':
            data = [
                {
                    'Numéro': c.get('numero'),
                    'Libellé': c.get('libelle'),
                    'Débit': c.get('debit', 0),
                    'Crédit': c.get('credit', 0),
                    'Solde': c.get('solde', 0),
                }
                for c in comptes
            ]
            return cls.generate_csv_response(
                data, headers,
                f"balance_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}"
            )

        elif format_type == 'excel':
            data = [
                {
                    'Numéro': c.get('numero'),
                    'Libellé': c.get('libelle'),
                    'Débit': c.get('debit', 0),
                    'Crédit': c.get('credit', 0),
                    'Solde': c.get('solde', 0),
                }
                for c in comptes
            ]
            return cls.generate_excel_response(
                data, headers,
                f"balance_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}",
                sheet_name='Balance'
            )

        else:  # PDF
            def build_balance_pdf(p):
                width, height = A4

                # En-tête
                p.setFont("Helvetica-Bold", 16)
                p.drawString(2*cm, height - 2*cm, titre)

                p.setFont("Helvetica", 10)
                p.drawString(2*cm, height - 2.7*cm,
                             f"Période: {date_debut.strftime('%d.%m.%Y')} - {date_fin.strftime('%d.%m.%Y')}")

                # Tableau
                y = height - 4*cm

                # En-têtes tableau
                p.setFont("Helvetica-Bold", 10)
                p.drawString(2*cm, y, "Numéro")
                p.drawString(4*cm, y, "Libellé")
                p.drawRightString(13*cm, y, "Débit")
                p.drawRightString(16*cm, y, "Crédit")
                p.drawRightString(19*cm, y, "Solde")

                p.line(2*cm, y - 0.2*cm, 19*cm, y - 0.2*cm)
                y -= 0.7*cm

                # Données
                p.setFont("Helvetica", 9)
                total_debit = Decimal('0')
                total_credit = Decimal('0')

                for compte in comptes:
                    if y < 3*cm:
                        p.showPage()
                        y = height - 2*cm
                        p.setFont("Helvetica", 9)

                    p.drawString(2*cm, y, str(compte.get('numero', '')))
                    p.drawString(4*cm, y, str(compte.get('libelle', ''))[:40])
                    p.drawRightString(13*cm, y, f"{compte.get('debit', 0):,.2f}")
                    p.drawRightString(16*cm, y, f"{compte.get('credit', 0):,.2f}")
                    p.drawRightString(19*cm, y, f"{compte.get('solde', 0):,.2f}")

                    total_debit += Decimal(str(compte.get('debit', 0)))
                    total_credit += Decimal(str(compte.get('credit', 0)))
                    y -= 0.5*cm

                # Totaux
                y -= 0.3*cm
                p.line(2*cm, y, 19*cm, y)
                y -= 0.5*cm
                p.setFont("Helvetica-Bold", 10)
                p.drawString(4*cm, y, "TOTAUX")
                p.drawRightString(13*cm, y, f"{total_debit:,.2f}")
                p.drawRightString(16*cm, y, f"{total_credit:,.2f}")
                p.drawRightString(19*cm, y, f"{total_debit - total_credit:,.2f}")

            return cls.generate_pdf_streaming(
                build_balance_pdf,
                f"balance_{date_debut.strftime('%Y%m%d')}_{date_fin.strftime('%Y%m%d')}"
            )


class QRBillService:
    """
    Service pour la génération de QR-Bill suisses conformes.
    Utilise la librairie qrbill pour la génération.
    """

    @staticmethod
    def generate_qr_reference(creditor_id: str, invoice_id: str) -> str:
        """
        Génère une référence QR structurée selon norme suisse (27 chiffres).

        Args:
            creditor_id: Identifiant du créancier (6 chiffres)
            invoice_id: Identifiant de la facture

        Returns:
            str: Référence QR complète avec checksum
        """
        # ID créancier (6 digits)
        id_creancier = str(creditor_id)[:6].zfill(6)

        # ID facture (20 digits)
        id_facture = str(invoice_id).zfill(20)

        # Référence de base (26 digits)
        ref_base = id_creancier + id_facture

        # Calcul checksum modulo 10 récursif (norme suisse)
        checksum = QRBillService._calculate_mod10_checksum(ref_base)

        return ref_base + str(checksum)

    @staticmethod
    def _calculate_mod10_checksum(reference: str) -> int:
        """Calcule le checksum modulo 10 récursif selon norme suisse."""
        table = [0, 9, 4, 6, 8, 2, 7, 1, 3, 5]
        carry = 0
        for char in reference:
            carry = table[(carry + int(char)) % 10]
        return (10 - carry) % 10

    @classmethod
    def generate_qrbill_svg(
        cls,
        iban: str,
        creditor: Dict,
        amount: Decimal,
        currency: str = 'CHF',
        debtor: Optional[Dict] = None,
        reference: Optional[str] = None,
        additional_info: str = '',
        language: str = 'fr'
    ) -> bytes:
        """
        Génère un QR-Bill au format SVG.

        Args:
            iban: IBAN du compte (QR-IBAN ou IBAN standard)
            creditor: Dict avec name, street, pcode, city, country
            amount: Montant
            currency: Devise (CHF ou EUR)
            debtor: Dict avec name, street, pcode, city, country (optionnel)
            reference: Référence QR ou SCOR (optionnel)
            additional_info: Information additionnelle
            language: Langue (fr, de, it, en)

        Returns:
            bytes: Contenu SVG du QR-Bill
        """
        from qrbill import QRBill
        import tempfile

        # Créer le QR-Bill
        qr_bill = QRBill(
            account=iban,
            creditor=creditor,
            debtor=debtor,
            amount=str(amount),
            reference_number=reference,
            additional_information=additional_info,
            language=language,
        )

        # Générer le SVG dans un fichier temporaire
        with tempfile.NamedTemporaryFile(suffix='.svg', delete=False, mode='w') as temp_svg:
            qr_bill.as_svg(temp_svg.name)

            # Lire le contenu
            with open(temp_svg.name, 'rb') as f:
                svg_content = f.read()

        return svg_content

    @classmethod
    def generate_qrbill_pdf(
        cls,
        iban: str,
        creditor: Dict,
        amount: Decimal,
        currency: str = 'CHF',
        debtor: Optional[Dict] = None,
        reference: Optional[str] = None,
        additional_info: str = '',
        language: str = 'fr'
    ) -> bytes:
        """
        Génère un QR-Bill au format PDF.

        Returns:
            bytes: Contenu PDF du QR-Bill
        """
        from svglib.svglib import svg2rlg
        from reportlab.graphics import renderPDF
        import tempfile

        # D'abord générer le SVG
        svg_content = cls.generate_qrbill_svg(
            iban=iban,
            creditor=creditor,
            amount=amount,
            currency=currency,
            debtor=debtor,
            reference=reference,
            additional_info=additional_info,
            language=language
        )

        # Convertir SVG en PDF via svglib
        with tempfile.NamedTemporaryFile(suffix='.svg', delete=False, mode='wb') as temp_svg:
            temp_svg.write(svg_content)
            temp_svg.flush()

            drawing = svg2rlg(temp_svg.name)

            pdf_buffer = io.BytesIO()
            renderPDF.drawToFile(drawing, pdf_buffer)
            pdf_buffer.seek(0)

            return pdf_buffer.getvalue()

    @classmethod
    def embed_qrbill_in_invoice(
        cls,
        invoice_pdf_buffer: io.BytesIO,
        iban: str,
        creditor: Dict,
        amount: Decimal,
        debtor: Optional[Dict] = None,
        reference: Optional[str] = None,
        additional_info: str = ''
    ) -> io.BytesIO:
        """
        Ajoute un QR-Bill à un PDF de facture existant.

        Le QR-Bill est ajouté sur la dernière page en bas,
        ou sur une nouvelle page si pas assez d'espace.

        Args:
            invoice_pdf_buffer: Buffer contenant le PDF de facture
            ... autres params pour le QR-Bill

        Returns:
            io.BytesIO: Buffer avec le PDF combiné
        """
        from PyPDF2 import PdfReader, PdfWriter
        import tempfile

        # Générer le QR-Bill SVG
        svg_content = cls.generate_qrbill_svg(
            iban=iban,
            creditor=creditor,
            amount=amount,
            debtor=debtor,
            reference=reference,
            additional_info=additional_info
        )

        # Convertir SVG en PDF
        from svglib.svglib import svg2rlg
        from reportlab.graphics import renderPDF
        from reportlab.lib.pagesizes import A4

        with tempfile.NamedTemporaryFile(suffix='.svg', delete=False, mode='wb') as temp_svg:
            temp_svg.write(svg_content)
            temp_svg.flush()

            drawing = svg2rlg(temp_svg.name)

            # Créer un PDF avec le QR-Bill positionné en bas
            qr_buffer = io.BytesIO()
            c = canvas.Canvas(qr_buffer, pagesize=A4)
            width, height = A4

            # Le QR-Bill fait environ 105mm x 210mm
            # On le place en bas de page
            from reportlab.graphics import renderPDF as renderPDFGraphics
            renderPDFGraphics.draw(drawing, c, 0, 0)

            c.save()
            qr_buffer.seek(0)

        # Merger les PDFs
        invoice_pdf_buffer.seek(0)
        reader = PdfReader(invoice_pdf_buffer)
        qr_reader = PdfReader(qr_buffer)
        writer = PdfWriter()

        # Copier toutes les pages de la facture
        for page in reader.pages:
            writer.add_page(page)

        # Ajouter la page QR-Bill
        if qr_reader.pages:
            writer.add_page(qr_reader.pages[0])

        # Sauvegarder
        output_buffer = io.BytesIO()
        writer.write(output_buffer)
        output_buffer.seek(0)

        return output_buffer
