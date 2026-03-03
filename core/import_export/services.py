# core/import_export/services.py
"""
Service d'import/export centralisé.

Fournit une interface unifiée pour:
- Importer des données depuis CSV/Excel
- Exporter des données vers CSV/Excel
- Générer des templates vides
- Mode dry-run (simulation)
"""

import io
from typing import Any, Dict, List, Optional, Type, Union
from decimal import Decimal
from datetime import datetime, date

from django.http import HttpResponse, StreamingHttpResponse
from django.utils.translation import gettext_lazy as _, gettext
from django.conf import settings

from import_export import resources
from import_export.formats.base_formats import CSV, XLSX
from import_export.results import Result

import tablib
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Protection
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment


class ImportExportService:
    """
    Service centralisé pour l'import/export de données.

    Méthodes principales:
    - import_file(): Importer depuis un fichier
    - export_queryset(): Exporter un queryset
    - generate_template(): Générer un template vide
    """

    # Taille maximale des fichiers (par défaut 10MB)
    MAX_FILE_SIZE = getattr(settings, 'IMPORT_EXPORT_MAX_FILE_SIZE', 10 * 1024 * 1024)

    # Formats supportés
    FORMATS = {
        'csv': CSV,
        'xlsx': XLSX,
    }

    # =========================================================================
    # IMPORT
    # =========================================================================

    @classmethod
    def import_file(
        cls,
        file,
        resource: resources.ModelResource,
        dry_run: bool = False,
        **context
    ) -> Dict[str, Any]:
        """
        Importe des données depuis un fichier.

        Args:
            file: Fichier uploadé (UploadedFile)
            resource: Instance de Resource à utiliser
            dry_run: Si True, simule l'import sans modifier la DB
            **context: Contexte supplémentaire (user, mandat, ip_address, etc.)

        Returns:
            Dict avec:
            - success: bool
            - total_rows: int
            - created: int
            - updated: int
            - skipped: int
            - errors: List[Dict]
            - warnings: List[str]
        """
        # Vérifier la taille du fichier
        if hasattr(file, 'size') and file.size > cls.MAX_FILE_SIZE:
            return {
                'success': False,
                'error': gettext(
                    "Le fichier est trop volumineux. Taille maximale: {size}MB"
                ).format(size=cls.MAX_FILE_SIZE // (1024 * 1024)),
            }

        # Détecter le format
        format_type = cls._detect_format(file.name)
        if format_type is None:
            return {
                'success': False,
                'error': gettext(
                    "Format de fichier non supporté. Utilisez CSV ou Excel (.xlsx)"
                ),
            }

        try:
            # Lire le fichier
            dataset = cls._read_file(file, format_type)

            # Valider les en-têtes
            validation_result = cls._validate_headers(dataset, resource)
            if not validation_result['valid']:
                return {
                    'success': False,
                    'error': validation_result['error'],
                    'missing_columns': validation_result.get('missing_columns', []),
                }

            # Effectuer l'import
            result = resource.import_data(
                dataset,
                dry_run=dry_run,
                raise_errors=False,
                **context
            )

            # Formater le résultat
            return cls._format_import_result(result, dry_run)

        except Exception as e:
            return {
                'success': False,
                'error': str(e),
            }

    @classmethod
    def _detect_format(cls, filename: str) -> Optional[str]:
        """Détecte le format du fichier à partir de son extension."""
        if filename.lower().endswith('.csv'):
            return 'csv'
        elif filename.lower().endswith('.xlsx'):
            return 'xlsx'
        elif filename.lower().endswith('.xls'):
            return 'xlsx'  # Traiter comme xlsx
        return None

    @classmethod
    def _read_file(cls, file, format_type: str) -> tablib.Dataset:
        """Lit le contenu du fichier et retourne un Dataset."""
        content = file.read()

        if format_type == 'csv':
            # Détecter l'encodage et le délimiteur
            try:
                text = content.decode('utf-8-sig')  # Gérer BOM
            except UnicodeDecodeError:
                text = content.decode('latin-1')

            # Détecter le délimiteur
            delimiter = ';' if ';' in text[:1000] else ','

            dataset = tablib.Dataset()
            dataset.load(text, format='csv', delimiter=delimiter)

        elif format_type == 'xlsx':
            dataset = tablib.Dataset()
            dataset.load(content, format='xlsx')

        else:
            raise ValueError(f"Format non supporté: {format_type}")

        return dataset

    @classmethod
    def _validate_headers(
        cls,
        dataset: tablib.Dataset,
        resource: resources.ModelResource
    ) -> Dict[str, Any]:
        """Valide que les colonnes requises sont présentes."""
        if not dataset.headers:
            return {
                'valid': False,
                'error': gettext("Le fichier ne contient pas d'en-têtes."),
            }

        # Normaliser les en-têtes (minuscules, sans espaces)
        file_headers = {h.lower().strip() for h in dataset.headers if h}

        # Récupérer les champs requis de la Resource
        required_fields = []
        optional_fields = []

        for field_name, field in resource.fields.items():
            if field_name in resource._meta.exclude:
                continue

            column_name = field.column_name or field_name
            normalized = column_name.lower().strip()

            # Un champ est requis s'il n'a pas de valeur par défaut
            # et n'est pas nullable
            model_field = cls._get_model_field(resource, field_name)
            if model_field and not model_field.null and not model_field.blank:
                if not hasattr(model_field, 'default') or model_field.default is None:
                    required_fields.append(normalized)
            else:
                optional_fields.append(normalized)

        # Vérifier les champs requis
        missing = [f for f in required_fields if f not in file_headers]

        if missing:
            return {
                'valid': False,
                'error': gettext(
                    "Colonnes requises manquantes: {columns}"
                ).format(columns=', '.join(missing)),
                'missing_columns': missing,
            }

        return {'valid': True}

    @staticmethod
    def _get_model_field(resource: resources.ModelResource, field_name: str):
        """Récupère le champ du modèle."""
        try:
            return resource._meta.model._meta.get_field(field_name)
        except Exception:
            return None

    @classmethod
    def _format_import_result(cls, result: Result, dry_run: bool) -> Dict[str, Any]:
        """Formate le résultat de l'import pour l'API."""
        from import_export.results import RowResult

        created = 0
        updated = 0
        skipped = 0
        errors = []

        for row in result.rows:
            if row.import_type == RowResult.IMPORT_TYPE_NEW:
                created += 1
            elif row.import_type == RowResult.IMPORT_TYPE_UPDATE:
                updated += 1
            elif row.import_type == RowResult.IMPORT_TYPE_SKIP:
                skipped += 1

            if row.errors:
                errors.append({
                    'row': row.number,
                    'errors': [str(e.error) for e in row.errors],
                })

        return {
            'success': not result.has_errors(),
            'dry_run': dry_run,
            'total_rows': result.total_rows,
            'created': created,
            'updated': updated,
            'skipped': skipped,
            'error_count': len(errors),
            'errors': errors[:50],  # Limiter à 50 erreurs
            'message': cls._get_result_message(created, updated, skipped, len(errors), dry_run),
        }

    @staticmethod
    def _get_result_message(
        created: int,
        updated: int,
        skipped: int,
        errors: int,
        dry_run: bool
    ) -> str:
        """Génère un message de résultat lisible."""
        if dry_run:
            prefix = gettext("Simulation: ")
        else:
            prefix = ""

        parts = []
        if created:
            parts.append(gettext("{n} créé(s)").format(n=created))
        if updated:
            parts.append(gettext("{n} mis à jour").format(n=updated))
        if skipped:
            parts.append(gettext("{n} ignoré(s)").format(n=skipped))
        if errors:
            parts.append(gettext("{n} erreur(s)").format(n=errors))

        if not parts:
            return prefix + gettext("Aucune donnée importée.")

        return prefix + ', '.join(parts) + '.'

    # =========================================================================
    # EXPORT
    # =========================================================================

    @classmethod
    def export_queryset(
        cls,
        queryset,
        resource: resources.ModelResource,
        format_type: str = 'xlsx',
        filename: str = 'export',
    ) -> HttpResponse:
        """
        Exporte un queryset vers un fichier téléchargeable.

        Args:
            queryset: QuerySet Django à exporter
            resource: Instance de Resource à utiliser
            format_type: 'csv' ou 'xlsx'
            filename: Nom du fichier (sans extension)

        Returns:
            HttpResponse avec le fichier à télécharger
        """
        # Exporter avec la Resource
        dataset = resource.export(queryset)

        # Générer le fichier
        if format_type == 'xlsx':
            return cls._export_excel(dataset, filename)
        else:
            return cls._export_csv(dataset, filename)

    @classmethod
    def _export_csv(cls, dataset: tablib.Dataset, filename: str) -> HttpResponse:
        """Exporte en CSV avec BOM UTF-8 pour Excel."""
        content = dataset.export('csv', delimiter=';')

        response = HttpResponse(
            '\ufeff' + content,  # BOM UTF-8
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response

    @classmethod
    def _export_excel(cls, dataset: tablib.Dataset, filename: str) -> HttpResponse:
        """Exporte en Excel avec mise en forme."""
        # Créer le fichier Excel avec openpyxl pour le styling
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Export'

        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-têtes
        if dataset.headers:
            for col, header in enumerate(dataset.headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

        # Données
        for row_idx, row in enumerate(dataset, 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)

                # Convertir les types
                if isinstance(value, Decimal):
                    cell.value = float(value)
                    cell.number_format = '#,##0.00'
                elif isinstance(value, (datetime, date)):
                    cell.value = value
                    cell.number_format = 'DD.MM.YYYY'
                else:
                    cell.value = value

                cell.border = thin_border

        # Auto-ajuster les colonnes
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Figer la première ligne
        ws.freeze_panes = 'A2'

        # Sauvegarder
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    # =========================================================================
    # TEMPLATE GENERATION
    # =========================================================================

    @classmethod
    def generate_template(
        cls,
        resource: resources.ModelResource,
        format_type: str = 'xlsx',
        filename: str = 'template',
        include_example: bool = True,
    ) -> HttpResponse:
        """
        Génère un template vide à remplir par l'utilisateur.

        Args:
            resource: Instance de Resource
            format_type: 'csv' ou 'xlsx'
            filename: Nom du fichier
            include_example: Si True, inclut une ligne d'exemple

        Returns:
            HttpResponse avec le fichier template
        """
        template_data = resource.get_template_data()

        if format_type == 'xlsx':
            return cls._generate_excel_template(template_data, filename, include_example)
        else:
            return cls._generate_csv_template(template_data, filename, include_example)

    @classmethod
    def _generate_csv_template(
        cls,
        template_data: Dict[str, Any],
        filename: str,
        include_example: bool
    ) -> HttpResponse:
        """Génère un template CSV."""
        headers = template_data['headers']
        example_row = template_data.get('example_row', {})

        lines = [';'.join(headers)]

        if include_example and example_row:
            values = [str(example_row.get(h, '')) for h in headers]
            lines.append(';'.join(values))

        content = '\n'.join(lines)

        response = HttpResponse(
            '\ufeff' + content,
            content_type='text/csv; charset=utf-8'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response

    @classmethod
    def _generate_excel_template(
        cls,
        template_data: Dict[str, Any],
        filename: str,
        include_example: bool
    ) -> HttpResponse:
        """
        Génère un template Excel enrichi avec:
        - En-têtes colorés
        - Descriptions des colonnes en commentaires
        - Ligne d'exemple
        - Validation des données si possible
        """
        headers = template_data['headers']
        descriptions = template_data.get('descriptions', {})
        example_row = template_data.get('example_row', {})

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template'

        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        example_font = Font(italic=True, color="666666")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # En-têtes avec commentaires
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

            # Ajouter un commentaire avec la description
            if header in descriptions:
                comment = Comment(
                    descriptions[header],
                    "AltiusOne"
                )
                comment.width = 300
                comment.height = 100
                cell.comment = comment

        # Ligne d'exemple
        if include_example and example_row:
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=example_row.get(header, ''))
                cell.font = example_font
                cell.fill = example_fill
                cell.border = thin_border

        # Ajuster la largeur des colonnes
        for col, header in enumerate(headers, 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = max(len(header) + 5, 15)

        # Figer la première ligne
        ws.freeze_panes = 'A2'

        # Protéger les en-têtes (optionnel)
        ws.protection.sheet = False
        for col in range(1, len(headers) + 1):
            ws.cell(row=1, column=col).protection = Protection(locked=True)

        # Instructions dans une deuxième feuille
        ws_help = wb.create_sheet(title='Instructions')
        ws_help['A1'] = gettext("Instructions d'import")
        ws_help['A1'].font = Font(bold=True, size=14)

        instructions = [
            "",
            gettext("1. Remplissez les données dans l'onglet 'Template'"),
            gettext("2. La première ligne (verte) est un exemple - vous pouvez la supprimer"),
            gettext("3. Ne modifiez pas les en-têtes de colonnes"),
            gettext("4. Utilisez les formats suivants:"),
            gettext("   - Dates: JJ.MM.AAAA (ex: 31.12.2024)"),
            gettext("   - Nombres: 1234.56 (point comme séparateur décimal)"),
            gettext("   - Booléens: Oui/Non"),
            "",
            gettext("5. Survolez les en-têtes pour voir les descriptions des champs"),
            "",
            gettext("Colonnes disponibles:"),
        ]

        for row, text in enumerate(instructions, 2):
            ws_help.cell(row=row, column=1, value=text)

        # Liste des colonnes avec descriptions
        start_row = len(instructions) + 3
        for idx, header in enumerate(headers):
            desc = descriptions.get(header, '')
            ws_help.cell(row=start_row + idx, column=1, value=f"• {header}")
            ws_help.cell(row=start_row + idx, column=2, value=desc)

        ws_help.column_dimensions['A'].width = 30
        ws_help.column_dimensions['B'].width = 60

        # Sauvegarder
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
        return response

    # =========================================================================
    # UTILITAIRES
    # =========================================================================

    @classmethod
    def get_available_resources(cls) -> Dict[str, Type[resources.ModelResource]]:
        """
        Retourne la liste des Resources disponibles.

        Returns:
            Dict avec app_label.model_name -> Resource class
        """
        # Import dynamique pour éviter les imports circulaires
        from core.import_export.resources import RESOURCE_REGISTRY
        return RESOURCE_REGISTRY
