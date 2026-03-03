# analytics/tasks.py
"""
Tâches Celery pour la génération asynchrone de rapports.
"""
import io
import time
import logging
from datetime import date
from decimal import Decimal

from celery import shared_task
from django.core.files.base import ContentFile
from django.utils import timezone
from django.utils.translation import gettext as _
from django.apps import apps

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm, mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, Image
)
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)

# Couleurs AltiusOne pour les graphiques
CHART_COLORS = ['#4680FF', '#2CA87F', '#DC2626', '#E58A00', '#673AB7', '#00BCD4', '#8BC34A', '#FF5722']
CHART_PRIMARY = '#366092'
CHART_SUCCESS = '#2CA87F'
CHART_DANGER = '#DC2626'
CHART_WARNING = '#E58A00'


def _hex_to_rgba(hex_color: str, alpha: float = 1.0) -> str:
    """Convertit une couleur hex (#RRGGBB) en rgba."""
    hex_color = hex_color.lstrip('#')
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    return f'rgba({r}, {g}, {b}, {alpha})'


def _generate_chart_image(chart_type: str, data: dict, title: str = '', width: float = 16, height: float = 10) -> bytes:
    """
    Génère une image de graphique avec Plotly (léger et non-bloquant).

    Args:
        chart_type: Type de graphique ('donut', 'bar', 'horizontal_bar', 'line', 'area')
        data: Données du graphique (labels, values, series, categories)
        title: Titre du graphique
        width: Largeur en cm
        height: Hauteur en cm

    Returns:
        bytes: Image PNG du graphique
    """
    import plotly.graph_objects as go

    # Convertir cm en pixels (environ 37.8 pixels par cm)
    width_px = int(width * 37.8)
    height_px = int(height * 37.8)

    # Créer la figure selon le type de graphique
    if chart_type == 'donut':
        fig = _create_donut_chart(data, title)
    elif chart_type == 'bar':
        fig = _create_bar_chart(data, title)
    elif chart_type == 'horizontal_bar':
        fig = _create_horizontal_bar_chart(data, title)
    elif chart_type == 'line':
        fig = _create_line_chart(data, title)
    elif chart_type == 'area':
        fig = _create_area_chart(data, title)
    elif chart_type == 'stacked_bar':
        fig = _create_stacked_bar_chart(data, title)
    else:
        # Par défaut: bar chart
        fig = _create_bar_chart(data, title)

    # Appliquer le layout commun
    fig.update_layout(
        width=width_px,
        height=height_px,
        margin=dict(l=50, r=50, t=50, b=50),
        paper_bgcolor='white',
        plot_bgcolor='white',
        font=dict(family='Arial, sans-serif', size=10),
    )

    # Exporter en PNG
    return fig.to_image(format='png', scale=2)


def _create_donut_chart(data: dict, title: str):
    """Crée un graphique en donut avec Plotly."""
    import plotly.graph_objects as go

    labels = data.get('labels', [])
    values = data.get('series', data.get('values', []))

    if not values or all(v == 0 for v in values):
        # Graphique vide avec message
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14))
        return fig

    # Filtrer les valeurs nulles
    filtered_data = [(l, v) for l, v in zip(labels, values) if v > 0]
    if filtered_data:
        labels, values = zip(*filtered_data)

    total = sum(values)

    fig = go.Figure(data=[go.Pie(
        labels=labels,
        values=values,
        hole=0.5,
        marker=dict(colors=CHART_COLORS[:len(values)]),
        textinfo='percent',
        textposition='inside',
        insidetextorientation='radial',
        hovertemplate='%{label}<br>CHF %{value:,.0f}<br>%{percent}<extra></extra>'
    )])

    # Ajouter le total au centre
    fig.add_annotation(
        text=f"CHF {total:,.0f}".replace(',', "'"),
        x=0.5, y=0.5,
        font=dict(size=14, family='Arial', color='black'),
        showarrow=False
    )

    if title:
        fig.update_layout(title=dict(text=title, x=0.5, font=dict(size=12)))

    return fig


def _create_bar_chart(data: dict, title: str):
    """Crée un graphique en barres verticales avec Plotly."""
    import plotly.graph_objects as go

    categories = data.get('categories', [])
    series = data.get('series', [])

    if not series or not categories:
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14))
        return fig

    fig = go.Figure()

    for i, s in enumerate(series):
        values = s.get('data', [])
        name = s.get('name', f'Série {i+1}')
        color = CHART_COLORS[i % len(CHART_COLORS)]

        fig.add_trace(go.Bar(
            x=categories,
            y=values,
            name=name,
            marker_color=color,
            hovertemplate='%{x}<br>CHF %{y:,.0f}<extra></extra>'
        ))

    fig.update_layout(
        barmode='group',
        xaxis=dict(tickangle=-45),
        yaxis=dict(gridcolor='rgba(0,0,0,0.1)', tickformat=',.0f'),
        showlegend=len(series) > 1,
    )

    if title:
        fig.update_layout(title=dict(text=title, x=0.5, font=dict(size=12)))

    return fig


def _create_horizontal_bar_chart(data: dict, title: str):
    """Crée un graphique en barres horizontales avec Plotly."""
    import plotly.graph_objects as go

    categories = data.get('categories', [])
    series = data.get('series', [])

    if not series or not categories:
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14))
        return fig

    values = series[0].get('data', []) if series else []
    colors_list = [CHART_SUCCESS if v >= 0 else CHART_DANGER for v in values]

    fig = go.Figure(data=[go.Bar(
        y=categories,
        x=values,
        orientation='h',
        marker_color=colors_list,
        hovertemplate='%{y}<br>CHF %{x:,.0f}<extra></extra>'
    )])

    fig.update_layout(
        xaxis=dict(gridcolor='rgba(0,0,0,0.1)', tickformat=',.0f'),
        yaxis=dict(autorange='reversed'),
    )

    if title:
        fig.update_layout(title=dict(text=title, x=0.5, font=dict(size=12)))

    return fig


def _create_line_chart(data: dict, title: str):
    """Crée un graphique en lignes avec Plotly."""
    import plotly.graph_objects as go

    categories = data.get('categories', [])
    series = data.get('series', [])

    if not series or not categories:
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14))
        return fig

    fig = go.Figure()

    for i, s in enumerate(series):
        values = s.get('data', [])
        name = s.get('name', f'Série {i+1}')
        color = CHART_COLORS[i % len(CHART_COLORS)]

        fig.add_trace(go.Scatter(
            x=categories,
            y=values,
            mode='lines+markers',
            name=name,
            line=dict(color=color, width=2),
            marker=dict(size=6),
            hovertemplate='%{x}<br>CHF %{y:,.0f}<extra></extra>'
        ))

    fig.update_layout(
        xaxis=dict(tickangle=-45, gridcolor='rgba(0,0,0,0.1)'),
        yaxis=dict(gridcolor='rgba(0,0,0,0.1)', tickformat=',.0f'),
        showlegend=len(series) > 1,
    )

    if title:
        fig.update_layout(title=dict(text=title, x=0.5, font=dict(size=12)))

    return fig


def _create_area_chart(data: dict, title: str):
    """Crée un graphique en aires avec Plotly."""
    import plotly.graph_objects as go

    categories = data.get('categories', [])
    series = data.get('series', [])

    if not series or not categories:
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14))
        return fig

    fig = go.Figure()

    for i, s in enumerate(series):
        values = s.get('data', [])
        name = s.get('name', f'Série {i+1}')
        color = CHART_COLORS[i % len(CHART_COLORS)]

        # Convertir la couleur hex en rgba pour le remplissage
        fill_color = _hex_to_rgba(color, 0.3) if color.startswith('#') else color.replace(')', ', 0.3)').replace('rgb', 'rgba')

        fig.add_trace(go.Scatter(
            x=categories,
            y=values,
            mode='lines',
            name=name,
            fill='tozeroy',
            line=dict(color=color, width=2),
            fillcolor=fill_color,
            hovertemplate='%{x}<br>CHF %{y:,.0f}<extra></extra>'
        ))

    fig.update_layout(
        xaxis=dict(tickangle=-45, gridcolor='rgba(0,0,0,0.1)'),
        yaxis=dict(gridcolor='rgba(0,0,0,0.1)', tickformat=',.0f'),
        showlegend=len(series) > 1,
    )

    if title:
        fig.update_layout(title=dict(text=title, x=0.5, font=dict(size=12)))

    return fig


def _create_stacked_bar_chart(data: dict, title: str):
    """Crée un graphique en barres empilées avec Plotly."""
    import plotly.graph_objects as go

    categories = data.get('categories', [])
    series = data.get('series', [])

    if not series or not categories:
        fig = go.Figure()
        fig.add_annotation(text="Aucune donnée", x=0.5, y=0.5, showarrow=False, font=dict(size=14))
        return fig

    fig = go.Figure()

    for i, s in enumerate(series):
        values = s.get('data', [])
        name = s.get('name', f'Série {i+1}')
        color = CHART_COLORS[i % len(CHART_COLORS)]

        fig.add_trace(go.Bar(
            x=categories,
            y=values,
            name=name,
            marker_color=color,
            hovertemplate='%{x}<br>%{fullData.name}: CHF %{y:,.0f}<extra></extra>'
        ))

    fig.update_layout(
        barmode='stack',
        xaxis=dict(tickangle=-45),
        yaxis=dict(gridcolor='rgba(0,0,0,0.1)', tickformat=',.0f'),
        showlegend=True,
    )

    if title:
        fig.update_layout(title=dict(text=title, x=0.5, font=dict(size=12)))

    return fig


def _create_chart_image_element(chart_type: str, data: dict, title: str = '', width: float = 14):
    """
    Crée un élément Image ReportLab à partir d'un graphique Plotly.

    Args:
        chart_type: Type de graphique
        data: Données du graphique
        title: Titre
        width: Largeur en cm

    Returns:
        Image: Element ReportLab ou None si erreur
    """
    try:
        # Vérifier que les données sont valides
        if not data:
            logger.warning(f"Graphique {chart_type}: données vides")
            return None

        # Vérifier la présence de series ou labels non vides
        series = data.get('series', [])
        labels = data.get('labels', [])
        categories = data.get('categories', [])

        # Pour les donuts/pies - utilise 'values' pas 'series'
        if chart_type in ['donut', 'pie']:
            values = data.get('values', [])
            if not values or all(v == 0 for v in values if isinstance(v, (int, float))):
                logger.info(f"Graphique {chart_type}: values vides ou nulles")
                return None
        # Pour les bar/area/line
        elif chart_type in ['bar', 'horizontal_bar', 'area', 'line', 'stacked_bar']:
            if not series:
                logger.info(f"Graphique {chart_type}: pas de series")
                return None
            has_data = False
            for s in series:
                if isinstance(s, dict) and s.get('data'):
                    if any(v != 0 for v in s['data'] if isinstance(v, (int, float))):
                        has_data = True
                        break
            if not has_data:
                logger.info(f"Graphique {chart_type}: toutes les données sont à 0")
                return None

        logger.info(f"Génération graphique {chart_type}: {title}")
        chart_bytes = _generate_chart_image(chart_type, data, title, width=width, height=width*0.6)

        if not chart_bytes:
            logger.error(f"Graphique {chart_type}: aucune donnée retournée par Plotly")
            return None

        chart_buffer = io.BytesIO(chart_bytes)

        # Créer l'élément Image ReportLab
        img = Image(chart_buffer, width=width*cm, height=width*0.6*cm)
        logger.info(f"Graphique {chart_type} créé avec succès ({len(chart_bytes)} bytes)")
        return img
    except ImportError as e:
        logger.error(f"Dépendance manquante pour graphique {chart_type}: {e}. Installez plotly et kaleido.")
        return None
    except Exception as e:
        logger.exception(f"Erreur création graphique {chart_type}: {e}")
        return None


def generer_pdf_preview(rapport) -> bytes:
    """
    Génère un PDF de preview à partir des sections du rapport.

    Cette fonction génère un PDF complet basé sur les sections configurées
    par l'utilisateur, permettant un aperçu fidèle du résultat final.

    Args:
        rapport: Instance du modèle Rapport

    Returns:
        bytes: Contenu du PDF ou None en cas d'erreur
    """
    from analytics.models import SectionRapport
    from analytics.services import RapportSectionService, GraphiqueService
    from html.parser import HTMLParser
    from io import StringIO

    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        styles = getSampleStyleSheet()

        # Styles personnalisés
        styles.add(ParagraphStyle(
            name='TitreRapport',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER
        ))
        styles.add(ParagraphStyle(
            name='SousTitre',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#366092')
        ))
        styles.add(ParagraphStyle(
            name='Info',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey
        ))
        styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
        ))

        elements = []

        # Récupérer les sections du rapport
        sections = RapportSectionService.get_sections_rapport(rapport)

        if not sections.exists():
            # Fallback: générer avec l'ancien système
            logger.info(f"Aucune section pour rapport {rapport.id}, utilisation du système legacy")
            content, _, _ = _generer_pdf(rapport)
            return content

        # Générer chaque section
        for section in sections:
            if not section.visible:
                continue

            section_elements = _generer_section_pdf(
                section=section,
                rapport=rapport,
                styles=styles,
            )
            elements.extend(section_elements)

        # Si aucun élément généré, ajouter un message
        if not elements:
            elements.append(Paragraph(
                "Aucun contenu à afficher. Ajoutez des sections au rapport.",
                styles['Normal']
            ))

        doc.build(elements)

        pdf_content = buffer.getvalue()
        buffer.close()

        logger.info(f"Preview PDF généré pour rapport {rapport.id}: {len(pdf_content)} bytes")
        return pdf_content

    except Exception as e:
        logger.exception(f"Erreur génération preview PDF: {e}")
        return None


def generer_pdf_preview_live(
    type_rapport: str,
    date_debut: str,
    date_fin: str,
    mandat=None,
    sections: list = None,
    options: dict = None
) -> bytes:
    """
    Génère un PDF de preview en temps réel sans rapport existant.

    Cette fonction permet de prévisualiser le rapport avant sa création.

    Args:
        type_rapport: Type de rapport (BILAN, COMPTE_RESULTATS, etc.)
        date_debut: Date de début (format YYYY-MM-DD)
        date_fin: Date de fin (format YYYY-MM-DD)
        mandat: Instance du modèle Mandat (optionnel)
        sections: Liste des sections configurées par l'utilisateur
        options: Options supplémentaires (inclure_comparatif, detail_comptes, etc.)

    Returns:
        bytes: Contenu du PDF ou None en cas d'erreur
    """
    from datetime import datetime as dt
    from analytics.models import TypeGraphiqueRapport
    from analytics.services import GraphiqueService

    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=2*cm,
            leftMargin=2*cm,
            topMargin=2*cm,
            bottomMargin=2*cm
        )

        styles = getSampleStyleSheet()

        # Styles personnalisés
        styles.add(ParagraphStyle(
            name='TitreRapport',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20,
            alignment=TA_CENTER
        ))
        styles.add(ParagraphStyle(
            name='SousTitre',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.HexColor('#366092')
        ))
        styles.add(ParagraphStyle(
            name='Info',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
            alignment=TA_CENTER
        ))
        styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=styles['Heading2'],
            fontSize=14,
            spaceBefore=15,
            spaceAfter=10,
        ))
        styles.add(ParagraphStyle(
            name='KPITitle',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey
        ))
        styles.add(ParagraphStyle(
            name='KPIValue',
            parent=styles['Normal'],
            fontSize=16,
            fontName='Helvetica-Bold'
        ))

        elements = []
        sections = sections or []
        options = options or {}

        # Parser les dates
        try:
            date_debut_dt = dt.strptime(date_debut, '%Y-%m-%d').date()
            date_fin_dt = dt.strptime(date_fin, '%Y-%m-%d').date()
        except ValueError:
            date_debut_dt = dt.now().date()
            date_fin_dt = dt.now().date()

        # En-tête du rapport (toujours présent)
        if mandat:
            elements.append(Paragraph(mandat.entreprise.nom if hasattr(mandat, 'entreprise') else str(mandat), styles['Info']))
            elements.append(Spacer(1, 5))

        # Si aucune section, générer un contenu par défaut
        if not sections:
            # Titre par défaut
            type_labels = {
                'BILAN': 'Bilan',
                'COMPTE_RESULTATS': 'Compte de résultats',
                'BALANCE': 'Balance',
                'TRESORERIE': 'Trésorerie',
                'TVA': 'Déclaration TVA',
                'SALAIRES': 'Salaires',
                'EVOLUTION_CA': "Évolution du chiffre d'affaires",
                'RENTABILITE': 'Rentabilité',
            }
            titre = type_labels.get(type_rapport, type_rapport)
            elements.append(Paragraph(titre, styles['TitreRapport']))
            elements.append(Paragraph(
                f"Période: {date_debut_dt.strftime('%d.%m.%Y')} au {date_fin_dt.strftime('%d.%m.%Y')}",
                styles['Info']
            ))
            elements.append(Spacer(1, 20))

            # Message indicatif
            elements.append(Paragraph(
                "Ajoutez des sections pour personnaliser le contenu du rapport.",
                styles['Normal']
            ))
        else:
            # Générer chaque section
            for section_data in sections:
                if not section_data.get('visible', True):
                    continue

                section_type = section_data.get('type_section', '')

                try:
                    if section_type == 'titre':
                        texte = _html_to_text(section_data.get('contenu_texte', ''))
                        if texte:
                            elements.append(Paragraph(texte, styles['TitreRapport']))

                    elif section_type == 'texte':
                        texte = _html_to_text(section_data.get('contenu_texte', ''))
                        if texte:
                            elements.append(Paragraph(texte, styles['Normal']))
                            elements.append(Spacer(1, 10))

                    elif section_type == 'graphique':
                        config = section_data.get('config', {})
                        code_graphique = config.get('code_graphique')
                        type_graphique_data = section_data.get('type_graphique', {})
                        chart_name = type_graphique_data.get('nom', code_graphique) if type_graphique_data else code_graphique

                        if code_graphique:
                            chart_img = None
                            try:
                                type_graphique = TypeGraphiqueRapport.objects.get(code=code_graphique)
                                chart_name = type_graphique.nom
                                logger.info(f"Génération graphique {code_graphique} (type={type_graphique.type_graphique})")

                                # Toujours essayer de récupérer les données
                                data = GraphiqueService.get_donnees_graphique(
                                    type_graphique=type_graphique,
                                    mandat=mandat,
                                    date_debut=date_debut_dt,
                                    date_fin=date_fin_dt,
                                )
                                logger.info(f"Données pour {code_graphique}: categories={len(data.get('categories', []))}, series={len(data.get('series', []))}, labels={len(data.get('labels', []))}, values={len(data.get('values', []))}")

                                titre = config.get('titre_personnalise', type_graphique.nom)
                                chart_img = _create_chart_image_element(
                                    type_graphique.type_graphique,
                                    data,
                                    titre,
                                    width=14
                                )
                            except TypeGraphiqueRapport.DoesNotExist:
                                logger.warning(f"Type graphique non trouvé: {code_graphique}")
                            except Exception as e:
                                logger.exception(f"Erreur génération graphique {code_graphique}: {e}")

                            if chart_img:
                                elements.append(chart_img)
                                elements.append(Spacer(1, 15))
                            else:
                                # Afficher un message informatif si pas de données
                                elements.append(Paragraph(chart_name or "Graphique", styles['SectionTitle']))
                                elements.append(Paragraph(
                                    "Aucune donnée disponible pour ce graphique sur la période sélectionnée.",
                                    ParagraphStyle('NoData', parent=styles['Normal'], textColor=colors.grey, fontSize=10, alignment=TA_CENTER)
                                ))
                                elements.append(Spacer(1, 15))

                    elif section_type == 'tableau':
                        config = section_data.get('config', {})
                        titre = config.get('titre', 'Détail des données')
                        max_lignes = config.get('max_lignes', 50)
                        elements.append(Paragraph(titre, styles['SectionTitle']))

                        # Récupérer les vraies données comptables
                        table_elements = _generer_tableau_preview(
                            type_rapport, mandat, date_debut_dt, date_fin_dt, styles, max_lignes
                        )
                        if table_elements:
                            elements.extend(table_elements)
                        else:
                            elements.append(Paragraph(
                                "Aucune donnée disponible pour la période sélectionnée.",
                                ParagraphStyle('NoData', parent=styles['Normal'], textColor=colors.grey, fontSize=10, alignment=TA_CENTER)
                            ))
                        elements.append(Spacer(1, 15))

                    elif section_type == 'kpi':
                        elements.append(Paragraph("Indicateurs clés", styles['SectionTitle']))
                        # Récupérer les vraies données KPI
                        kpi_elements = _generer_kpis_preview(
                            type_rapport, mandat, date_debut_dt, date_fin_dt, styles
                        )
                        if kpi_elements:
                            elements.extend(kpi_elements)
                        else:
                            elements.append(Paragraph(
                                "Aucun indicateur disponible pour la période sélectionnée.",
                                ParagraphStyle('NoData', parent=styles['Normal'], textColor=colors.grey, fontSize=10, alignment=TA_CENTER)
                            ))
                        elements.append(Spacer(1, 15))

                    elif section_type == 'saut_page':
                        elements.append(PageBreak())

                    elif section_type == 'separateur':
                        from reportlab.platypus import HRFlowable
                        elements.append(Spacer(1, 10))
                        elements.append(HRFlowable(
                            width="100%",
                            thickness=1,
                            color=colors.lightgrey,
                            spaceBefore=5,
                            spaceAfter=10
                        ))

                except Exception as e:
                    logger.warning(f"Erreur section preview {section_type}: {e}")
                    elements.append(Paragraph(
                        f"[Erreur: {section_type}]",
                        ParagraphStyle('Error', parent=styles['Normal'], textColor=colors.red)
                    ))

        # Footer avec date de génération
        elements.append(Spacer(1, 30))
        elements.append(Paragraph(
            f"Aperçu généré le {dt.now().strftime('%d.%m.%Y à %H:%M')}",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        ))

        doc.build(elements)

        pdf_content = buffer.getvalue()
        buffer.close()

        logger.info(f"Preview PDF live généré: {len(pdf_content)} bytes")
        return pdf_content

    except Exception as e:
        logger.exception(f"Erreur génération preview PDF live: {e}")
        return None


def _generer_tableau_preview(type_rapport: str, mandat, date_debut, date_fin, styles, max_lignes: int = 50) -> list:
    """
    Génère le tableau de données pour un preview (sans rapport existant).
    """
    elements = []

    if not mandat:
        return elements

    # Créer un objet temporaire pour utiliser _get_donnees_comptables
    class TempRapport:
        def __init__(self, type_r, m, d1, d2):
            self.type_rapport = type_r
            self.mandat = m
            self.date_debut = d1
            self.date_fin = d2

    temp_rapport = TempRapport(type_rapport, mandat, date_debut, date_fin)
    data = _get_donnees_comptables(temp_rapport, type_rapport.lower())

    if type_rapport == 'BILAN':
        if data.get('actif'):
            elements.append(Paragraph("ACTIF", styles['Heading3']))
            actif_data = [['Compte', 'Libellé', 'Montant']]
            for item in data['actif'][:max_lignes]:
                actif_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])
            if len(actif_data) > 1:
                table = Table(actif_data, colWidths=[3*cm, 10*cm, 4*cm])
                table.setStyle(_get_table_style())
                elements.append(table)
            elements.append(Spacer(1, 10))

        if data.get('passif'):
            elements.append(Paragraph("PASSIF", styles['Heading3']))
            passif_data = [['Compte', 'Libellé', 'Montant']]
            for item in data['passif'][:max_lignes]:
                passif_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])
            if len(passif_data) > 1:
                table = Table(passif_data, colWidths=[3*cm, 10*cm, 4*cm])
                table.setStyle(_get_table_style())
                elements.append(table)

    elif type_rapport == 'COMPTE_RESULTATS':
        if data.get('produits'):
            elements.append(Paragraph("PRODUITS", styles['Heading3']))
            table_data = [['Compte', 'Libellé', 'Montant']]
            for item in data['produits'][:max_lignes]:
                table_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])
            if len(table_data) > 1:
                table = Table(table_data, colWidths=[3*cm, 10*cm, 4*cm])
                table.setStyle(_get_table_style())
                elements.append(table)
            elements.append(Spacer(1, 10))

        if data.get('charges'):
            elements.append(Paragraph("CHARGES", styles['Heading3']))
            table_data = [['Compte', 'Libellé', 'Montant']]
            for item in data['charges'][:max_lignes]:
                table_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])
            if len(table_data) > 1:
                table = Table(table_data, colWidths=[3*cm, 10*cm, 4*cm])
                table.setStyle(_get_table_style())
                elements.append(table)

    else:
        # Générique
        for key, items in data.items():
            if isinstance(items, list) and items:
                elements.append(Paragraph(key.upper(), styles['Heading3']))
                table_data = [['Compte', 'Libellé', 'Montant']]
                for item in items[:max_lignes]:
                    if isinstance(item, dict):
                        table_data.append([
                            item.get('numero', ''),
                            item.get('libelle', ''),
                            _format_montant(item.get('solde', 0))
                        ])
                if len(table_data) > 1:
                    table = Table(table_data, colWidths=[3*cm, 10*cm, 4*cm])
                    table.setStyle(_get_table_style())
                    elements.append(table)
                elements.append(Spacer(1, 10))

    return elements


def _generer_kpis_preview(type_rapport: str, mandat, date_debut, date_fin, styles) -> list:
    """
    Génère les indicateurs clés pour un preview (sans rapport existant).
    """
    elements = []

    if not mandat:
        return elements

    # Créer un objet temporaire
    class TempRapport:
        def __init__(self, type_r, m, d1, d2):
            self.type_rapport = type_r
            self.mandat = m
            self.date_debut = d1
            self.date_fin = d2

    temp_rapport = TempRapport(type_rapport, mandat, date_debut, date_fin)
    data = _get_donnees_comptables(temp_rapport, type_rapport.lower())

    kpis = []

    if type_rapport == 'BILAN':
        total_actif = sum(float(item['solde']) for item in data.get('actif', []))
        total_passif = sum(float(item['solde']) for item in data.get('passif', []))
        kpis = [
            ('Total Actif', _format_montant(total_actif)),
            ('Total Passif', _format_montant(total_passif)),
        ]

    elif type_rapport == 'COMPTE_RESULTATS':
        total_produits = sum(float(item['solde']) for item in data.get('produits', []))
        total_charges = sum(float(item['solde']) for item in data.get('charges', []))
        resultat = total_produits - total_charges
        kpis = [
            ('Total Produits', _format_montant(total_produits)),
            ('Total Charges', _format_montant(total_charges)),
            ('Résultat net', _format_montant(resultat)),
        ]

    if kpis:
        kpi_data = [[k, v] for k, v in kpis]
        kpi_table = Table(kpi_data, colWidths=[10*cm, 5*cm])
        kpi_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica'),
            ('FONTNAME', (1, 0), (1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(kpi_table)

    return elements


def _generer_section_pdf(section, rapport, styles) -> list:
    """
    Génère les éléments ReportLab pour une section.

    Args:
        section: Instance SectionRapport
        rapport: Instance Rapport
        styles: Styles ReportLab

    Returns:
        Liste d'éléments ReportLab
    """
    from analytics.services import RapportSectionService, GraphiqueService
    import re

    elements = []

    try:
        if section.type_section == 'titre':
            # Formater le contenu avec les variables
            contenu = RapportSectionService.formater_contenu_avec_variables(
                section.contenu_texte, rapport
            )
            # Extraire le texte du HTML
            texte = _html_to_text(contenu)
            elements.append(Paragraph(texte, styles['TitreRapport']))

        elif section.type_section == 'texte':
            contenu = RapportSectionService.formater_contenu_avec_variables(
                section.contenu_texte, rapport
            )
            # Convertir HTML simple en Paragraph
            # Pour l'instant, on strip le HTML et on utilise le texte brut
            # TODO: Support complet du HTML
            texte = _html_to_text(contenu)
            if texte:
                elements.append(Paragraph(texte, styles['Normal']))
                elements.append(Spacer(1, 10))

        elif section.type_section == 'graphique':
            if section.type_graphique:
                # Récupérer les données du graphique prédéfini
                data = GraphiqueService.get_donnees_graphique(
                    type_graphique=section.type_graphique,
                    mandat=rapport.mandat,
                    date_debut=rapport.date_debut,
                    date_fin=rapport.date_fin,
                )

                # Générer l'image du graphique
                titre = section.config.get('titre_personnalise', section.type_graphique.nom)
                chart_img = _create_chart_image_element(
                    section.type_graphique.type_graphique,
                    data,
                    titre,
                    width=14
                )

                if chart_img:
                    elements.append(chart_img)
                    elements.append(Spacer(1, 15))

        elif section.type_section == 'tableau':
            # Générer le tableau de données
            config = section.config or {}
            titre = config.get('titre', '')
            max_lignes = config.get('max_lignes', 50)

            if titre:
                elements.append(Paragraph(titre, styles['SectionTitle']))

            # Récupérer les données selon le type de rapport
            table_elements = _generer_tableau_section(rapport, styles, max_lignes)
            elements.extend(table_elements)

        elif section.type_section == 'kpi':
            # Générer les KPIs
            kpi_elements = _generer_kpis_section(rapport, styles)
            elements.extend(kpi_elements)

        elif section.type_section == 'saut_page':
            elements.append(PageBreak())

        elif section.type_section == 'separateur':
            # Ligne horizontale
            from reportlab.platypus import HRFlowable
            elements.append(Spacer(1, 10))
            elements.append(HRFlowable(
                width="100%",
                thickness=1,
                color=colors.lightgrey,
                spaceBefore=5,
                spaceAfter=10
            ))

    except Exception as e:
        logger.exception(f"Erreur génération section {section.type_section}: {e}")
        elements.append(Paragraph(
            f"[Erreur section {section.type_section}]",
            ParagraphStyle('Error', parent=styles['Normal'], textColor=colors.red)
        ))

    return elements


def _html_to_text(html: str) -> str:
    """Convertit du HTML simple en texte."""
    import re
    # Supprimer les balises
    text = re.sub(r'<[^>]+>', '', html)
    # Décoder les entités HTML basiques
    text = text.replace('&nbsp;', ' ')
    text = text.replace('&amp;', '&')
    text = text.replace('&lt;', '<')
    text = text.replace('&gt;', '>')
    text = text.replace('&quot;', '"')
    return text.strip()


def _generer_tableau_section(rapport, styles, max_lignes: int = 50) -> list:
    """
    Génère le tableau de données pour une section tableau.

    Le contenu dépend du type de rapport.
    """
    elements = []

    # Utiliser les fonctions existantes pour récupérer les données
    data = _get_donnees_comptables(rapport, rapport.type_rapport.lower())

    if rapport.type_rapport == 'BILAN':
        # Tableau actif
        if data.get('actif'):
            elements.append(Paragraph("ACTIF", styles['Heading3']))
            actif_data = [['Compte', 'Libellé', 'Montant']]
            for item in data['actif'][:max_lignes]:
                actif_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])
            if len(actif_data) > 1:
                table = Table(actif_data, colWidths=[3*cm, 10*cm, 4*cm])
                table.setStyle(_get_table_style())
                elements.append(table)
            elements.append(Spacer(1, 10))

        # Tableau passif
        if data.get('passif'):
            elements.append(Paragraph("PASSIF", styles['Heading3']))
            passif_data = [['Compte', 'Libellé', 'Montant']]
            for item in data['passif'][:max_lignes]:
                passif_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])
            if len(passif_data) > 1:
                table = Table(passif_data, colWidths=[3*cm, 10*cm, 4*cm])
                table.setStyle(_get_table_style())
                elements.append(table)

    elif rapport.type_rapport == 'COMPTE_RESULTATS':
        # Tableau produits
        if data.get('produits'):
            elements.append(Paragraph("PRODUITS", styles['Heading3']))
            table_data = [['Compte', 'Libellé', 'Montant']]
            for item in data['produits'][:max_lignes]:
                table_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])
            if len(table_data) > 1:
                table = Table(table_data, colWidths=[3*cm, 10*cm, 4*cm])
                table.setStyle(_get_table_style())
                elements.append(table)
            elements.append(Spacer(1, 10))

        # Tableau charges
        if data.get('charges'):
            elements.append(Paragraph("CHARGES", styles['Heading3']))
            table_data = [['Compte', 'Libellé', 'Montant']]
            for item in data['charges'][:max_lignes]:
                table_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])
            if len(table_data) > 1:
                table = Table(table_data, colWidths=[3*cm, 10*cm, 4*cm])
                table.setStyle(_get_table_style())
                elements.append(table)

    else:
        # Générique: afficher les données brutes si disponibles
        for key, items in data.items():
            if isinstance(items, list) and items:
                elements.append(Paragraph(key.upper(), styles['Heading3']))
                table_data = [['Compte', 'Libellé', 'Montant']]
                for item in items[:max_lignes]:
                    if isinstance(item, dict):
                        table_data.append([
                            item.get('numero', ''),
                            item.get('libelle', ''),
                            _format_montant(item.get('solde', 0))
                        ])
                if len(table_data) > 1:
                    table = Table(table_data, colWidths=[3*cm, 10*cm, 4*cm])
                    table.setStyle(_get_table_style())
                    elements.append(table)
                elements.append(Spacer(1, 10))

    return elements


def _generer_kpis_section(rapport, styles) -> list:
    """
    Génère les indicateurs clés (KPIs) pour une section.
    """
    elements = []

    # Récupérer les données
    data = _get_donnees_comptables(rapport, rapport.type_rapport.lower())

    kpis = []

    if rapport.type_rapport == 'BILAN':
        total_actif = sum(float(item['solde']) for item in data.get('actif', []))
        total_passif = sum(float(item['solde']) for item in data.get('passif', []))
        fonds_propres = total_passif  # Simplifié

        kpis = [
            ('Total Actif', _format_montant(total_actif)),
            ('Total Passif', _format_montant(total_passif)),
            ('Fonds propres', _format_montant(fonds_propres)),
        ]

    elif rapport.type_rapport == 'COMPTE_RESULTATS':
        total_produits = sum(float(item['solde']) for item in data.get('produits', []))
        total_charges = sum(float(item['solde']) for item in data.get('charges', []))
        resultat = total_produits - total_charges

        kpis = [
            ('Total Produits', _format_montant(total_produits)),
            ('Total Charges', _format_montant(total_charges)),
            ('Résultat net', _format_montant(resultat)),
        ]

    # Générer un mini tableau pour les KPIs
    if kpis:
        kpi_data = [[k, v] for k, v in kpis]
        kpi_table = Table(kpi_data, colWidths=[10*cm, 5*cm])
        kpi_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
        ]))
        elements.append(kpi_table)
        elements.append(Spacer(1, 15))

    return elements


@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def generer_rapport_async(self, rapport_id: str):
    """
    Génère un rapport de manière asynchrone.

    Args:
        rapport_id: UUID du rapport à générer
    """
    from analytics.models import Rapport
    from analytics.kpi_service import KPICalculator

    try:
        rapport = Rapport.objects.select_related('mandat', 'genere_par').get(id=rapport_id)
    except Rapport.DoesNotExist:
        logger.error(f"Rapport {rapport_id} non trouvé")
        return {'success': False, 'error': 'Rapport non trouvé'}

    if rapport.statut == 'TERMINE':
        logger.info(f"Rapport {rapport_id} déjà généré")
        return {'success': True, 'already_done': True}

    start_time = time.time()

    try:
        logger.info(f"Génération du rapport {rapport.nom} ({rapport.type_rapport})")

        # Générer le contenu selon le format
        if rapport.format_fichier == 'PDF':
            content, filename, nb_pages = _generer_pdf(rapport)
            rapport.nombre_pages = nb_pages
        elif rapport.format_fichier in ['EXCEL', 'XLSX']:
            content, filename = _generer_excel(rapport)
        elif rapport.format_fichier == 'CSV':
            content, filename = _generer_csv(rapport)
        elif rapport.format_fichier == 'HTML':
            content, filename = _generer_html(rapport)
        else:
            raise ValueError(f"Format non supporté: {rapport.format_fichier}")

        # Sauvegarder le fichier
        rapport.fichier.save(filename, ContentFile(content))
        rapport.taille_fichier = len(content)
        rapport.statut = 'TERMINE'
        rapport.duree_generation_secondes = int(time.time() - start_time)
        rapport.save()

        logger.info(f"Rapport {rapport.nom} généré avec succès en {rapport.duree_generation_secondes}s")

        # Envoyer par email si demandé
        if rapport.envoi_email and rapport.destinataires:
            _envoyer_rapport_email(rapport)

        return {
            'success': True,
            'rapport_id': str(rapport.id),
            'duree': rapport.duree_generation_secondes
        }

    except Exception as e:
        logger.exception(f"Erreur lors de la génération du rapport {rapport_id}")
        rapport.statut = 'ERREUR'
        rapport.parametres['erreur'] = str(e)
        rapport.save()

        # Retry si possible
        if self.request.retries < self.max_retries:
            raise self.retry(exc=e, countdown=60 * (self.request.retries + 1))

        return {'success': False, 'error': str(e)}


def _generer_pdf(rapport):
    """Génère le contenu PDF du rapport."""
    from analytics.kpi_service import KPICalculator

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )

    styles = getSampleStyleSheet()

    # Styles personnalisés
    styles.add(ParagraphStyle(
        name='TitreRapport',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        alignment=TA_CENTER
    ))
    styles.add(ParagraphStyle(
        name='SousTitre',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.HexColor('#366092')
    ))
    styles.add(ParagraphStyle(
        name='Info',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey
    ))
    styles.add(ParagraphStyle(
        name='SectionTitle',
        parent=styles['Heading2'],
        fontSize=14,
        spaceBefore=15,
        spaceAfter=10,
    ))

    elements = []

    # Vérifier si le rapport a des sections personnalisées
    sections = list(rapport.sections.filter(visible=True).order_by('ordre'))

    if sections:
        # Utiliser les sections personnalisées de l'utilisateur
        logger.info(f"Génération PDF avec {len(sections)} sections personnalisées")

        # Infos rapport en en-tête
        if rapport.mandat:
            elements.append(Paragraph(
                rapport.mandat.client.raison_sociale if hasattr(rapport.mandat, 'client') else str(rapport.mandat),
                styles['Info']
            ))
            elements.append(Spacer(1, 5))

        # Générer chaque section
        for section in sections:
            section_elements = _generer_section_pdf(section, rapport, styles)
            elements.extend(section_elements)

        # Footer
        elements.append(Spacer(1, 30))
        user = rapport.genere_par
        if user:
            nom_utilisateur = user.get_full_name() if hasattr(user, 'get_full_name') and user.get_full_name() else user.username
        else:
            nom_utilisateur = "Système"
        elements.append(Paragraph(
            f"Généré le {timezone.now().strftime('%d.%m.%Y à %H:%M')} par {nom_utilisateur}",
            ParagraphStyle('Footer', parent=styles['Normal'], fontSize=8, textColor=colors.grey, alignment=TA_CENTER)
        ))

    else:
        # Fallback: utiliser l'ancien système de génération par type
        logger.info(f"Génération PDF avec générateur par défaut pour {rapport.type_rapport}")

        # En-tête
        elements.append(Paragraph(rapport.nom, styles['TitreRapport']))

        # Infos rapport
        periode_text = f"Période: {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}"
        elements.append(Paragraph(periode_text, styles['Info']))

        if rapport.mandat:
            elements.append(Paragraph(f"Client: {rapport.mandat.client.raison_sociale}", styles['Info']))

        # Nom complet de l'utilisateur (prénom + nom, sinon username)
        user = rapport.genere_par
        if user:
            nom_utilisateur = user.get_full_name() if hasattr(user, 'get_full_name') else None
            if not nom_utilisateur:
                prenom = getattr(user, 'first_name', '') or ''
                nom = getattr(user, 'last_name', '') or ''
                nom_utilisateur = f"{prenom} {nom}".strip()
            if not nom_utilisateur:
                nom_utilisateur = user.username
        else:
            nom_utilisateur = "Système"
        elements.append(Paragraph(
            f"Généré le {timezone.now().strftime('%d.%m.%Y à %H:%M')} par {nom_utilisateur}",
            styles['Info']
        ))
        elements.append(Spacer(1, 20))

        # Contenu selon le type de rapport
        type_generators = {
            'BILAN': _generer_contenu_bilan,
            'COMPTE_RESULTATS': _generer_contenu_compte_resultats,
            'BALANCE': _generer_contenu_balance,
            'TRESORERIE': _generer_contenu_tresorerie,
            'TVA': _generer_contenu_tva,
            'SALAIRES': _generer_contenu_salaires,
            'EVOLUTION_CA': _generer_contenu_evolution_ca,
            'RENTABILITE': _generer_contenu_rentabilite,
            'CUSTOM': _generer_contenu_custom,
        }

        generator = type_generators.get(rapport.type_rapport, _generer_contenu_custom)
        content_elements = generator(rapport, styles)
        elements.extend(content_elements)

    # Pied de page avec numéro de page sera ajouté par le template
    doc.build(elements)

    pdf_content = buffer.getvalue()
    buffer.close()

    filename = f"{rapport.nom.replace(' ', '_')}_{rapport.date_fin.strftime('%Y%m%d')}.pdf"

    # Estimer le nombre de pages (approximatif)
    nb_pages = max(1, len(elements) // 20)

    return pdf_content, filename, nb_pages


def _generer_contenu_bilan(rapport, styles):
    """Génère le contenu du bilan avec graphique."""
    elements = []
    elements.append(Paragraph("Bilan", styles['SousTitre']))

    # Description du rapport
    elements.append(Paragraph(
        "Le bilan présente la situation financière de l'entreprise à une date donnée, "
        "en montrant d'un côté les actifs (ce que l'entreprise possède) et de l'autre "
        "les passifs (ce qu'elle doit).",
        styles['Normal']
    ))
    elements.append(Spacer(1, 15))

    # Récupérer les données comptables
    data = _get_donnees_comptables(rapport, 'bilan')

    # Vérifier si des données existent
    has_data = bool(data.get('actif') or data.get('passif'))

    if not has_data:
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            "⚠️ Aucune donnée comptable disponible pour la période sélectionnée.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange, fontSize=11)
        ))
        elements.append(Paragraph(
            f"Période: {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}",
            styles['Info']
        ))
        if rapport.mandat:
            elements.append(Paragraph(f"Mandat: {rapport.mandat.numero}", styles['Info']))
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            "Veuillez vérifier que des écritures comptables existent pour cette période et ce mandat.",
            styles['Normal']
        ))
        return elements

    # Calculer les totaux
    total_actif = sum(float(item['solde']) for item in data.get('actif', []))
    total_passif = sum(float(item['solde']) for item in data.get('passif', []))

    # === GRAPHIQUE DONUT ===
    chart_data = {
        'labels': ['Actif', 'Passif'],
        'series': [total_actif, total_passif]
    }
    chart_img = _create_chart_image_element('donut', chart_data, 'Répartition Actif / Passif', width=14)
    if chart_img:
        elements.append(chart_img)
        elements.append(Spacer(1, 20))

    # === TABLEAU ACTIF ===
    elements.append(Paragraph("ACTIF", styles['Heading3']))
    actif_data = [['Compte', 'Libellé', 'Montant']]
    for item in data.get('actif', []):
        actif_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])

    if len(actif_data) > 1:
        actif_table = Table(actif_data, colWidths=[3*cm, 10*cm, 4*cm])
        actif_table.setStyle(_get_table_style())
        elements.append(actif_table)
    else:
        elements.append(Paragraph("Aucune donnée disponible", styles['Normal']))

    elements.append(Spacer(1, 15))

    # === TABLEAU PASSIF ===
    elements.append(Paragraph("PASSIF", styles['Heading3']))
    passif_data = [['Compte', 'Libellé', 'Montant']]
    for item in data.get('passif', []):
        passif_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])

    if len(passif_data) > 1:
        passif_table = Table(passif_data, colWidths=[3*cm, 10*cm, 4*cm])
        passif_table.setStyle(_get_table_style())
        elements.append(passif_table)
    else:
        elements.append(Paragraph("Aucune donnée disponible", styles['Normal']))

    # === TOTAUX ===
    elements.append(Spacer(1, 20))

    totaux_data = [
        ['Total Actif', _format_montant(total_actif)],
        ['Total Passif', _format_montant(total_passif)],
    ]
    totaux_table = Table(totaux_data, colWidths=[13*cm, 4*cm])
    totaux_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(totaux_table)

    # Analyse
    elements.append(Spacer(1, 15))
    equilibre = total_actif - total_passif
    if abs(equilibre) < 0.01:
        elements.append(Paragraph(
            "✓ Le bilan est équilibré.",
            ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'))
        ))
    else:
        elements.append(Paragraph(
            f"⚠️ Écart de {_format_montant(abs(equilibre))} CHF entre l'actif et le passif.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange)
        ))

    return elements


def _generer_contenu_compte_resultats(rapport, styles):
    """Génère le contenu du compte de résultats avec graphique."""
    elements = []
    elements.append(Paragraph("Compte de résultats", styles['SousTitre']))

    # Description
    elements.append(Paragraph(
        "Le compte de résultats présente les performances financières de l'entreprise sur la période, "
        "en comparant les produits (revenus) aux charges (dépenses) pour déterminer le résultat net.",
        styles['Normal']
    ))
    elements.append(Spacer(1, 15))

    data = _get_donnees_comptables(rapport, 'resultats')

    # Calculer les totaux
    total_produits = sum(float(item['solde']) for item in data.get('produits', []))
    total_charges = sum(float(item['solde']) for item in data.get('charges', []))
    resultat = total_produits - total_charges

    # === GRAPHIQUE EN BARRES ===
    chart_data = {
        'categories': ['Produits', 'Charges', 'Résultat'],
        'series': [{
            'name': 'Montants',
            'data': [total_produits, total_charges, resultat]
        }]
    }
    chart_img = _create_chart_image_element('bar', chart_data, 'Synthèse du compte de résultats', width=14)
    if chart_img:
        elements.append(chart_img)
        elements.append(Spacer(1, 20))

    # === PRODUITS ===
    elements.append(Paragraph("PRODUITS", styles['Heading3']))
    produits_data = [['Compte', 'Libellé', 'Montant']]
    for item in data.get('produits', []):
        produits_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])

    if len(produits_data) > 1:
        table = Table(produits_data, colWidths=[3*cm, 10*cm, 4*cm])
        table.setStyle(_get_table_style())
        elements.append(table)
    else:
        elements.append(Paragraph("Aucun produit enregistré pour cette période", styles['Normal']))

    elements.append(Spacer(1, 15))

    # === CHARGES ===
    elements.append(Paragraph("CHARGES", styles['Heading3']))
    charges_data = [['Compte', 'Libellé', 'Montant']]
    for item in data.get('charges', []):
        charges_data.append([item['numero'], item['libelle'], _format_montant(item['solde'])])

    if len(charges_data) > 1:
        table = Table(charges_data, colWidths=[3*cm, 10*cm, 4*cm])
        table.setStyle(_get_table_style())
        elements.append(table)
    else:
        elements.append(Paragraph("Aucune charge enregistrée pour cette période", styles['Normal']))

    # === RÉSULTAT ===
    elements.append(Spacer(1, 20))

    resultat_data = [
        ['Total Produits', _format_montant(total_produits)],
        ['Total Charges', _format_montant(total_charges)],
        ['Résultat', _format_montant(resultat)],
    ]
    resultat_table = Table(resultat_data, colWidths=[13*cm, 4*cm])
    resultat_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEABOVE', (0, 2), (-1, 2), 1, colors.black),
        ('TEXTCOLOR', (1, 2), (1, 2), colors.HexColor('#2CA87F') if resultat >= 0 else colors.red),
    ]))
    elements.append(resultat_table)

    # Analyse du résultat
    elements.append(Spacer(1, 15))
    if resultat > 0:
        marge = (resultat / total_produits * 100) if total_produits > 0 else 0
        elements.append(Paragraph(
            f"✓ L'entreprise dégage un bénéfice de CHF {_format_montant(resultat)}, "
            f"soit une marge nette de {marge:.1f}%.",
            ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'))
        ))
    elif resultat < 0:
        elements.append(Paragraph(
            f"⚠️ L'entreprise enregistre une perte de CHF {_format_montant(abs(resultat))}.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.red)
        ))
    else:
        elements.append(Paragraph(
            "Le résultat est à l'équilibre (ni bénéfice, ni perte).",
            styles['Normal']
        ))

    return elements


def _generer_contenu_balance(rapport, styles):
    """Génère le contenu de la balance."""
    elements = []
    elements.append(Paragraph("Balance des comptes", styles['SousTitre']))

    data = _get_donnees_comptables(rapport, 'balance')

    table_data = [['N° Compte', 'Libellé', 'Débit', 'Crédit', 'Solde']]
    for item in data.get('comptes', []):
        table_data.append([
            item['numero'],
            item['libelle'],
            _format_montant(item.get('debit', 0)),
            _format_montant(item.get('credit', 0)),
            _format_montant(item.get('solde', 0))
        ])

    if len(table_data) > 1:
        table = Table(table_data, colWidths=[2.5*cm, 8*cm, 3*cm, 3*cm, 3*cm])
        table.setStyle(_get_table_style())
        elements.append(table)
    else:
        elements.append(Paragraph("Aucune donnée disponible pour cette période", styles['Normal']))

    return elements


def _generer_contenu_tresorerie(rapport, styles):
    """Génère le contenu du tableau de trésorerie avec graphiques."""
    elements = []
    elements.append(Paragraph("Tableau de trésorerie", styles['SousTitre']))

    # Description
    elements.append(Paragraph(
        "Le tableau de trésorerie présente les flux de liquidités de l'entreprise, "
        "montrant les entrées et sorties de fonds sur la période.",
        styles['Normal']
    ))
    elements.append(Spacer(1, 15))

    # Récupérer les vraies données
    data = _get_donnees_tresorerie(rapport)
    encaissements = data.get('encaissements', Decimal('0'))
    decaissements = data.get('decaissements', Decimal('0'))
    variation = data.get('variation', Decimal('0'))
    nb_factures_payees = data.get('nb_factures_payees', 0)
    delai_moyen_paiement = data.get('delai_moyen_paiement', 0)

    if encaissements == 0 and decaissements == 0:
        elements.append(Paragraph(
            "⚠️ Aucune donnée de trésorerie disponible pour cette période.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange)
        ))
        return elements

    # === GRAPHIQUE 1: FLUX DE TRÉSORERIE ===
    chart_data = {
        'categories': ['Encaissements', 'Décaissements', 'Solde net'],
        'series': [{
            'name': 'Montants',
            'data': [float(encaissements), float(decaissements), float(variation)]
        }]
    }
    chart_img = _create_chart_image_element('bar', chart_data, 'Flux de trésorerie', width=14)
    if chart_img:
        elements.append(chart_img)
        elements.append(Spacer(1, 15))

    # === GRAPHIQUE 2: RÉPARTITION ===
    if encaissements > 0 and decaissements > 0:
        chart_data_donut = {
            'labels': ['Encaissements', 'Décaissements'],
            'series': [float(encaissements), float(decaissements)]
        }
        chart_img2 = _create_chart_image_element('donut', chart_data_donut, 'Répartition des flux', width=12)
        if chart_img2:
            elements.append(chart_img2)
            elements.append(Spacer(1, 20))

    # === TABLEAU SYNTHÈSE ===
    elements.append(Paragraph("Synthèse des flux", styles['Heading3']))
    table_data = [
        ['Indicateur', 'Valeur'],
        ['Encaissements (factures payées)', f"CHF {_format_montant(encaissements)}"],
        ['Nombre de factures payées', str(nb_factures_payees)],
        ['Décaissements (estimé)', f"CHF {_format_montant(decaissements)}"],
        ['Solde net de trésorerie', f"CHF {_format_montant(variation)}"],
        ['Délai moyen de paiement', f"{delai_moyen_paiement} jours" if delai_moyen_paiement > 0 else 'N/A'],
    ]

    table = Table(table_data, colWidths=[10*cm, 7*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('FONTNAME', (0, -2), (-1, -2), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, -2), (1, -2),
            colors.HexColor('#2CA87F') if variation >= 0 else colors.red),
    ]))
    elements.append(table)

    # === ANALYSE ===
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("Analyse", styles['Heading3']))

    if variation > 0:
        elements.append(Paragraph(
            f"✓ Flux de trésorerie positif: +CHF {_format_montant(variation)} sur la période.",
            ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'))
        ))
    elif variation < 0:
        elements.append(Paragraph(
            f"⚠️ Flux de trésorerie négatif: CHF {_format_montant(variation)} sur la période.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.red)
        ))
    else:
        elements.append(Paragraph(
            "Les flux de trésorerie sont équilibrés sur la période.",
            styles['Normal']
        ))

    if delai_moyen_paiement > 0:
        if delai_moyen_paiement <= 30:
            elements.append(Paragraph(
                f"✓ Délai moyen de paiement excellent ({delai_moyen_paiement} jours).",
                ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'))
            ))
        elif delai_moyen_paiement <= 60:
            elements.append(Paragraph(
                f"◉ Délai moyen de paiement acceptable ({delai_moyen_paiement} jours).",
                styles['Normal']
            ))
        else:
            elements.append(Paragraph(
                f"⚠️ Délai moyen de paiement élevé ({delai_moyen_paiement} jours) - à surveiller.",
                ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange)
            ))

    return elements


def _generer_contenu_tva(rapport, styles):
    """Génère le contenu du rapport TVA avec graphique."""
    elements = []
    elements.append(Paragraph("Rapport TVA", styles['SousTitre']))

    # Description
    elements.append(Paragraph(
        "Ce rapport présente le décompte TVA pour la période sélectionnée, "
        "incluant la TVA due sur le chiffre d'affaires et l'impôt préalable déductible.",
        styles['Normal']
    ))
    elements.append(Paragraph(
        f"Période: {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}",
        styles['Info']
    ))
    elements.append(Spacer(1, 15))

    # Récupérer les données TVA
    data = _get_donnees_tva(rapport)
    tva_due = float(data.get('tva_due', 0))
    tva_deductible = float(data.get('tva_deductible', 0))
    tva_nette = float(data.get('tva_nette', 0))

    # === GRAPHIQUE DONUT ===
    if tva_due > 0 or tva_deductible > 0:
        chart_data = {
            'labels': ['TVA due', 'TVA déductible'],
            'series': [tva_due, tva_deductible]
        }
        chart_img = _create_chart_image_element('donut', chart_data, 'Répartition TVA', width=12)
        if chart_img:
            elements.append(chart_img)
            elements.append(Spacer(1, 20))

    # === TABLEAU ===
    elements.append(Paragraph("Décompte TVA", styles['Heading3']))
    table_data = [
        ['Description', 'Base', 'TVA'],
        ['Chiffre d\'affaires imposable', _format_montant(data.get('ca_imposable', 0)), ''],
        ['TVA due sur CA (taux normal)', '', _format_montant(data.get('tva_due', 0))],
        ['TVA due sur CA (taux réduit)', '', _format_montant(data.get('tva_due_reduit', 0))],
        ['', '', ''],
        ['Impôt préalable déductible', '', _format_montant(data.get('tva_deductible', 0))],
        ['', '', ''],
        ['TVA nette à payer', '', _format_montant(data.get('tva_nette', 0))],
    ]

    table = Table(table_data, colWidths=[10*cm, 4*cm, 3*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
        ('TEXTCOLOR', (2, -1), (2, -1), colors.HexColor('#2CA87F') if tva_nette <= 0 else colors.red),
    ]))
    elements.append(table)

    # === RÉSUMÉ ===
    elements.append(Spacer(1, 15))
    if tva_nette > 0:
        elements.append(Paragraph(
            f"💰 TVA nette à verser: CHF {_format_montant(tva_nette)}",
            ParagraphStyle('Info', parent=styles['Normal'], textColor=colors.HexColor('#366092'), fontSize=11)
        ))
    elif tva_nette < 0:
        elements.append(Paragraph(
            f"✓ Crédit de TVA à récupérer: CHF {_format_montant(abs(tva_nette))}",
            ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'), fontSize=11)
        ))
    else:
        elements.append(Paragraph(
            "La TVA est équilibrée (ni montant à payer, ni crédit).",
            styles['Normal']
        ))

    return elements


def _generer_contenu_salaires(rapport, styles):
    """Génère le contenu du rapport salaires avec graphique."""
    elements = []
    elements.append(Paragraph("Rapport Salaires", styles['SousTitre']))

    # Description
    elements.append(Paragraph(
        "Ce rapport présente le détail des salaires versés sur la période, "
        "incluant les salaires bruts, les charges sociales et les montants nets.",
        styles['Normal']
    ))
    elements.append(Spacer(1, 15))

    # Récupérer les données salaires
    data = _get_donnees_salaires(rapport)
    employes = data.get('employes', [])

    if not employes:
        elements.append(Paragraph(
            "⚠️ Aucune donnée salariale disponible pour cette période.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange)
        ))
        return elements

    total_brut = float(data.get('total_brut', 0))
    total_charges = float(data.get('total_charges', 0))
    total_net = float(data.get('total_net', 0))

    # === GRAPHIQUE DONUT (Répartition) ===
    chart_data = {
        'labels': ['Salaires nets', 'Charges sociales'],
        'series': [total_net, total_charges]
    }
    chart_img = _create_chart_image_element('donut', chart_data, 'Répartition masse salariale', width=12)
    if chart_img:
        elements.append(chart_img)
        elements.append(Spacer(1, 20))

    # === TABLEAU DÉTAILLÉ ===
    elements.append(Paragraph("Détail par employé", styles['Heading3']))
    table_data = [['Employé', 'Salaire brut', 'Charges sociales', 'Net à payer']]
    for emp in employes:
        table_data.append([
            emp['nom'],
            _format_montant(emp['brut']),
            _format_montant(emp['charges']),
            _format_montant(emp['net'])
        ])

    # Totaux
    table_data.append([
        'TOTAL',
        _format_montant(data.get('total_brut', 0)),
        _format_montant(data.get('total_charges', 0)),
        _format_montant(data.get('total_net', 0))
    ])

    table = Table(table_data, colWidths=[6*cm, 4*cm, 4*cm, 4*cm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f5f5f5')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e0e0e0')),
    ]))
    elements.append(table)

    # === STATISTIQUES ===
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("Statistiques", styles['Heading3']))

    nb_employes = len(employes)
    salaire_moyen = total_brut / nb_employes if nb_employes > 0 else 0
    taux_charges = (total_charges / total_brut * 100) if total_brut > 0 else 0

    stats_data = [
        ['Nombre d\'employés', str(nb_employes)],
        ['Salaire brut moyen', f"CHF {_format_montant(salaire_moyen)}"],
        ['Taux de charges sociales', f"{taux_charges:.1f}%"],
        ['Coût total employeur', f"CHF {_format_montant(total_brut)}"],
    ]
    stats_table = Table(stats_data, colWidths=[10*cm, 7*cm])
    stats_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(stats_table)

    return elements


def _generer_contenu_evolution_ca(rapport, styles):
    """Génère le contenu du rapport d'évolution du CA avec graphique."""
    elements = []
    elements.append(Paragraph("Évolution du Chiffre d'Affaires", styles['SousTitre']))

    # Description
    elements.append(Paragraph(
        "Ce rapport présente l'évolution mensuelle du chiffre d'affaires sur la période sélectionnée, "
        "permettant d'identifier les tendances et les variations saisonnières.",
        styles['Normal']
    ))
    elements.append(Spacer(1, 15))

    # Données mensuelles
    data = _get_donnees_ca_mensuel(rapport)
    mois_list = data.get('mois', [])

    if not mois_list:
        elements.append(Paragraph(
            "⚠️ Aucune donnée de chiffre d'affaires disponible pour cette période.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange)
        ))
        return elements

    # === GRAPHIQUE EN AIRES ===
    chart_data = {
        'categories': [m['libelle'] for m in mois_list],
        'series': [{
            'name': 'CA HT',
            'data': [float(m['ca']) for m in mois_list]
        }]
    }
    chart_img = _create_chart_image_element('area', chart_data, 'Évolution mensuelle du CA', width=15)
    if chart_img:
        elements.append(chart_img)
        elements.append(Spacer(1, 20))

    # === TABLEAU ===
    table_data = [['Mois', 'CA HT', 'Évolution']]
    for mois in mois_list:
        evol = f"+{mois['evolution']}%" if mois['evolution'] >= 0 else f"{mois['evolution']}%"
        table_data.append([mois['libelle'], _format_montant(mois['ca']), evol])

    if len(table_data) > 1:
        table = Table(table_data, colWidths=[6*cm, 6*cm, 5*cm])
        table.setStyle(_get_table_style())
        elements.append(table)

    # === ANALYSE ===
    elements.append(Spacer(1, 15))
    total = sum(float(m['ca']) for m in mois_list)
    moyenne = total / len(mois_list) if mois_list else 0
    ca_values = [float(m['ca']) for m in mois_list]
    ca_max = max(ca_values) if ca_values else 0
    ca_min = min(ca_values) if ca_values else 0
    mois_max = mois_list[ca_values.index(ca_max)]['libelle'] if ca_values else ''
    mois_min = mois_list[ca_values.index(ca_min)]['libelle'] if ca_values else ''

    # Calculer la tendance
    if len(ca_values) >= 2:
        premiere_moitie = sum(ca_values[:len(ca_values)//2]) / (len(ca_values)//2)
        deuxieme_moitie = sum(ca_values[len(ca_values)//2:]) / (len(ca_values) - len(ca_values)//2)
        tendance = ((deuxieme_moitie - premiere_moitie) / premiere_moitie * 100) if premiere_moitie > 0 else 0
    else:
        tendance = 0

    elements.append(Paragraph("Analyse", styles['Heading3']))
    analysis_data = [
        ['Total période', f"CHF {_format_montant(total)}"],
        ['Moyenne mensuelle', f"CHF {_format_montant(moyenne)}"],
        ['Meilleur mois', f"{mois_max} ({_format_montant(ca_max)} CHF)"],
        ['Mois le plus faible', f"{mois_min} ({_format_montant(ca_min)} CHF)"],
    ]
    analysis_table = Table(analysis_data, colWidths=[8*cm, 9*cm])
    analysis_table.setStyle(TableStyle([
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
    ]))
    elements.append(analysis_table)

    elements.append(Spacer(1, 10))
    if tendance > 5:
        elements.append(Paragraph(
            f"📈 Tendance haussière: +{tendance:.1f}% sur la période.",
            ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'))
        ))
    elif tendance < -5:
        elements.append(Paragraph(
            f"📉 Tendance baissière: {tendance:.1f}% sur la période.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.red)
        ))
    else:
        elements.append(Paragraph(
            "📊 Chiffre d'affaires stable sur la période.",
            styles['Normal']
        ))

    return elements


def _generer_contenu_rentabilite(rapport, styles):
    """Génère le contenu du rapport de rentabilité avec graphiques."""
    elements = []
    elements.append(Paragraph("Analyse de Rentabilité", styles['SousTitre']))

    # Description
    elements.append(Paragraph(
        "Ce rapport analyse les indicateurs clés de rentabilité de l'entreprise, "
        "permettant d'évaluer l'efficacité opérationnelle et la performance financière.",
        styles['Normal']
    ))
    elements.append(Spacer(1, 15))

    data = _get_donnees_rentabilite(rapport)

    # Vérifier si des données existent
    ca_total = data.get('ca_total', 0)
    if ca_total == 0:
        elements.append(Paragraph(
            "⚠️ Aucune donnée de chiffre d'affaires disponible pour cette période.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange)
        ))
        return elements

    # === GRAPHIQUE 1: RÉPARTITION CA / CHARGES / MARGE ===
    chart_data_repartition = {
        'labels': ['Chiffre d\'affaires', 'Charges', 'Marge brute'],
        'series': [
            data.get('ca_total', 0),
            data.get('charges_total', 0),
            data.get('marge_brute_montant', 0)
        ]
    }
    chart_img1 = _create_chart_image_element('donut', chart_data_repartition, 'Répartition financière', width=12)
    if chart_img1:
        elements.append(chart_img1)
        elements.append(Spacer(1, 15))

    # === GRAPHIQUE 2: INDICATEURS EN POURCENTAGE ===
    chart_data_pct = {
        'categories': ['Marge brute', 'Taux de rentabilité', 'Taux de facturation'],
        'series': [{
            'name': 'Pourcentage',
            'data': [
                data.get('marge_brute', 0),
                data.get('taux_rentabilite', 0),
                data.get('taux_facturation', 0)
            ]
        }]
    }
    chart_img2 = _create_chart_image_element('horizontal_bar', chart_data_pct, 'Indicateurs de rentabilité (%)', width=14)
    if chart_img2:
        elements.append(chart_img2)
        elements.append(Spacer(1, 20))

    # === SYNTHÈSE FINANCIÈRE ===
    elements.append(Paragraph("Synthèse financière", styles['Heading3']))
    synthese_data = [
        ['Indicateur', 'Montant CHF'],
        ['Chiffre d\'affaires HT', _format_montant(data.get('ca_total', 0))],
        ['Total des charges', _format_montant(data.get('charges_total', 0))],
        ['Marge brute', _format_montant(data.get('marge_brute_montant', 0))],
        ['Résultat net', _format_montant(data.get('resultat_net', 0))],
    ]

    synthese_table = Table(synthese_data, colWidths=[10*cm, 7*cm])
    synthese_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('TEXTCOLOR', (1, -1), (1, -1),
            colors.HexColor('#2CA87F') if data.get('resultat_net', 0) >= 0 else colors.red),
    ]))
    elements.append(synthese_table)
    elements.append(Spacer(1, 15))

    # === KPIs ===
    elements.append(Paragraph("Indicateurs clés de performance", styles['Heading3']))
    kpi_data = [
        ['Indicateur', 'Valeur', 'Évolution'],
        ['Marge brute', f"{data.get('marge_brute', 0):.1f}%", data.get('evol_marge', '+0%')],
        ['Taux de rentabilité', f"{data.get('taux_rentabilite', 0):.1f}%", data.get('evol_rentabilite', '+0%')],
        ['Heures facturées (estimé)', f"{data.get('heures_facturees', 0)}h", ''],
        ['Taux de facturation', f"{data.get('taux_facturation', 0):.1f}%", ''],
    ]

    table = Table(kpi_data, colWidths=[8*cm, 5*cm, 4*cm])
    table.setStyle(_get_table_style())
    elements.append(table)

    # === ANALYSE ===
    elements.append(Spacer(1, 15))
    elements.append(Paragraph("Analyse", styles['Heading3']))

    marge = data.get('marge_brute', 0)
    taux_rent = data.get('taux_rentabilite', 0)
    taux_fact = data.get('taux_facturation', 0)
    resultat = data.get('resultat_net', 0)

    # Analyse du résultat
    if resultat > 0:
        elements.append(Paragraph(
            f"✓ Résultat positif de CHF {_format_montant(resultat)} sur la période.",
            ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'))
        ))
    elif resultat < 0:
        elements.append(Paragraph(
            f"⚠️ Résultat négatif de CHF {_format_montant(abs(resultat))} sur la période.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.red)
        ))

    # Analyse de la marge
    if marge >= 30:
        elements.append(Paragraph(
            f"✓ Marge brute excellente ({marge:.1f}%) - La structure de coûts est bien maîtrisée.",
            ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'))
        ))
    elif marge >= 20:
        elements.append(Paragraph(
            f"◉ Marge brute correcte ({marge:.1f}%) - Des améliorations sont possibles.",
            styles['Normal']
        ))
    elif marge > 0:
        elements.append(Paragraph(
            f"⚠️ Marge brute faible ({marge:.1f}%) - Attention à la rentabilité.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange)
        ))

    # Analyse du taux de facturation
    if taux_fact >= 75:
        elements.append(Paragraph(
            f"✓ Taux de facturation optimal ({taux_fact:.1f}%) - Bonne utilisation des ressources.",
            ParagraphStyle('Success', parent=styles['Normal'], textColor=colors.HexColor('#2CA87F'))
        ))
    elif taux_fact >= 60:
        elements.append(Paragraph(
            f"◉ Taux de facturation acceptable ({taux_fact:.1f}%) - Potentiel d'amélioration.",
            styles['Normal']
        ))
    elif taux_fact > 0:
        elements.append(Paragraph(
            f"⚠️ Taux de facturation bas ({taux_fact:.1f}%) - Optimiser l'allocation des ressources.",
            ParagraphStyle('Warning', parent=styles['Normal'], textColor=colors.orange)
        ))

    return elements


def _generer_contenu_custom(rapport, styles):
    """Génère le contenu d'un rapport personnalisé."""
    elements = []
    elements.append(Paragraph("Rapport personnalisé", styles['SousTitre']))

    # Afficher les paramètres du rapport
    if rapport.parametres:
        elements.append(Paragraph("Configuration du rapport:", styles['Heading3']))
        for key, value in rapport.parametres.items():
            if key != 'erreur':
                elements.append(Paragraph(f"• {key}: {value}", styles['Normal']))

    elements.append(Spacer(1, 20))
    elements.append(Paragraph(
        "Ce rapport personnalisé nécessite une configuration spécifique. "
        "Veuillez contacter l'administrateur pour définir le contenu.",
        styles['Normal']
    ))

    return elements


def _get_table_style():
    """Retourne le style de tableau standard."""
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
    ])


def _format_montant(montant):
    """Formate un montant en CHF."""
    if montant is None:
        return '-'
    if isinstance(montant, Decimal):
        montant = float(montant)
    return f"{montant:,.2f}".replace(',', "'")


def _get_donnees_comptables(rapport, type_donnees):
    """Récupère les données comptables pour le rapport."""
    try:
        Compte = apps.get_model('comptabilite', 'Compte')
        EcritureComptable = apps.get_model('comptabilite', 'EcritureComptable')

        # Filtre de base
        filters = {
            'date_ecriture__gte': rapport.date_debut,
            'date_ecriture__lte': rapport.date_fin,
        }
        if rapport.mandat:
            filters['mandat'] = rapport.mandat

        from django.db.models import Sum

        # Récupérer les soldes par compte
        ecritures = EcritureComptable.objects.filter(**filters).values(
            'compte__numero', 'compte__libelle', 'compte__classe'
        ).annotate(
            total_debit=Sum('montant_debit'),
            total_credit=Sum('montant_credit')
        )

        result = {'actif': [], 'passif': [], 'produits': [], 'charges': [], 'comptes': []}

        for e in ecritures:
            solde = (e['total_debit'] or Decimal('0')) - (e['total_credit'] or Decimal('0'))
            item = {
                'numero': e['compte__numero'],
                'libelle': e['compte__libelle'],
                'debit': e['total_debit'] or Decimal('0'),
                'credit': e['total_credit'] or Decimal('0'),
                'solde': abs(solde)
            }

            classe = e['compte__classe']

            # Classification selon le plan comptable suisse PME (classe est un IntegerField)
            # Classe 1 = Actifs
            # Classe 2 = Passifs (y compris capitaux propres)
            # Classe 3 = Produits d'exploitation (ventes, prestations)
            # Classe 4 = Charges de marchandises et prestations de tiers
            # Classe 5 = Charges de personnel
            # Classe 6 = Autres charges d'exploitation
            # Classe 7 = Produits hors exploitation
            # Classe 8 = Charges hors exploitation
            # Classe 9 = Clôture
            if classe == 1:  # Actif
                result['actif'].append(item)
            elif classe == 2:  # Passif (y compris capitaux propres)
                result['passif'].append(item)
            elif classe in [3, 7]:  # Produits (exploitation + hors exploitation)
                result['produits'].append(item)
            elif classe in [4, 5, 6, 8]:  # Charges (toutes les catégories)
                result['charges'].append(item)

            result['comptes'].append(item)

        return result

    except Exception as e:
        logger.warning(f"Erreur récupération données comptables: {e}")
        return {'actif': [], 'passif': [], 'produits': [], 'charges': [], 'comptes': []}


def _get_donnees_tresorerie(rapport):
    """Récupère les données de trésorerie pour le rapport."""
    try:
        Facture = apps.get_model('facturation', 'Facture')
        Paiement = apps.get_model('facturation', 'Paiement')

        from django.db.models import Sum
        from django.db.models.functions import Coalesce

        # Récupérer les paiements sur la période (table Paiement)
        paiement_filters = {
            'date_paiement__gte': rapport.date_debut,
            'date_paiement__lte': rapport.date_fin,
        }
        if rapport.mandat:
            paiement_filters['facture__mandat'] = rapport.mandat

        paiements = Paiement.objects.filter(**paiement_filters)
        encaissements = paiements.aggregate(
            total=Coalesce(Sum('montant'), Decimal('0'))
        )['total'] or Decimal('0')

        # Calcul du délai moyen de paiement
        delai_moyen_paiement = 0
        paiements_avec_dates = paiements.select_related('facture').filter(
            facture__date_emission__isnull=False
        )[:100]
        if paiements_avec_dates.exists():
            delais = []
            for p in paiements_avec_dates:
                if p.date_paiement and p.facture.date_emission:
                    delai = (p.date_paiement - p.facture.date_emission).days
                    if delai >= 0:
                        delais.append(delai)
            delai_moyen_paiement = int(sum(delais) / len(delais)) if delais else 0

        # Compter les factures payées (utiliser date_paiement_complet)
        facture_filters = {
            'statut__in': ['PAYEE', 'PARTIELLEMENT_PAYEE'],
            'date_paiement_complet__gte': rapport.date_debut,
            'date_paiement_complet__lte': rapport.date_fin,
        }
        if rapport.mandat:
            facture_filters['mandat'] = rapport.mandat
        nb_factures_payees = Facture.objects.filter(**facture_filters).count()

        # Si pas de paiements, fallback sur montant_paye des factures
        if encaissements == 0:
            factures_avec_paiement = Facture.objects.filter(
                statut__in=['PAYEE', 'PARTIELLEMENT_PAYEE'],
                montant_paye__gt=0,
                date_emission__gte=rapport.date_debut,
                date_emission__lte=rapport.date_fin,
            )
            if rapport.mandat:
                factures_avec_paiement = factures_avec_paiement.filter(mandat=rapport.mandat)

            encaissements = factures_avec_paiement.aggregate(
                total=Coalesce(Sum('montant_paye'), Decimal('0'))
            )['total'] or Decimal('0')
            nb_factures_payees = factures_avec_paiement.count()

        # Décaissements (charges comptables)
        try:
            EcritureComptable = apps.get_model('comptabilite', 'EcritureComptable')
            ecriture_filters = {
                'date_ecriture__gte': rapport.date_debut,
                'date_ecriture__lte': rapport.date_fin,
                'compte__classe__in': [4, 5, 6, 8],  # Charges selon plan comptable suisse PME
            }
            if rapport.mandat:
                ecriture_filters['mandat'] = rapport.mandat

            charges_data = EcritureComptable.objects.filter(**ecriture_filters).aggregate(
                total_debit=Sum('montant_debit'),
                total_credit=Sum('montant_credit')
            )
            decaissements = (charges_data['total_debit'] or Decimal('0')) - (charges_data['total_credit'] or Decimal('0'))
        except Exception:
            decaissements = encaissements * Decimal('0.7')

        if decaissements <= 0:
            decaissements = encaissements * Decimal('0.7')

        variation = encaissements - decaissements

        return {
            'encaissements': encaissements,
            'decaissements': decaissements,
            'variation': variation,
            'nb_factures_payees': nb_factures_payees,
            'delai_moyen_paiement': delai_moyen_paiement,
        }

    except Exception as e:
        logger.warning(f"Erreur récupération données trésorerie: {e}")
        return {
            'encaissements': Decimal('0'),
            'decaissements': Decimal('0'),
            'variation': Decimal('0'),
            'nb_factures_payees': 0,
            'delai_moyen_paiement': 0,
        }


def _get_donnees_tva(rapport):
    """Récupère les données TVA pour le rapport."""
    try:
        OperationTVA = apps.get_model('tva', 'OperationTVA')

        from django.db.models import Sum

        filters = {
            'date_operation__gte': rapport.date_debut,
            'date_operation__lte': rapport.date_fin,
        }
        if rapport.mandat:
            filters['mandat'] = rapport.mandat

        operations = OperationTVA.objects.filter(**filters).aggregate(
            ca_imposable=Sum('base_imposable'),
            tva_due=Sum('montant_tva'),
        )

        # Calculer TVA déductible (simplification)
        tva_due = operations['tva_due'] or Decimal('0')
        tva_deductible = tva_due * Decimal('0.6')  # Estimation

        return {
            'ca_imposable': operations['ca_imposable'] or Decimal('0'),
            'tva_due': tva_due,
            'tva_due_reduit': Decimal('0'),
            'tva_deductible': tva_deductible,
            'tva_nette': tva_due - tva_deductible,
        }

    except Exception as e:
        logger.warning(f"Erreur récupération données TVA: {e}")
        return {
            'ca_imposable': Decimal('0'),
            'tva_due': Decimal('0'),
            'tva_due_reduit': Decimal('0'),
            'tva_deductible': Decimal('0'),
            'tva_nette': Decimal('0'),
        }


def _get_donnees_salaires(rapport):
    """Récupère les données salaires pour le rapport."""
    try:
        Employe = apps.get_model('salaires', 'Employe')
        FicheSalaire = apps.get_model('salaires', 'FicheSalaire')

        from django.db.models import Sum, Q

        # Le modèle FicheSalaire utilise 'mois' et 'annee' au lieu de date_debut/date_fin
        filters = Q()
        if rapport.date_debut and rapport.date_fin:
            filters &= Q(annee__gte=rapport.date_debut.year, annee__lte=rapport.date_fin.year)
            if rapport.date_debut.year == rapport.date_fin.year:
                filters &= Q(mois__gte=rapport.date_debut.month, mois__lte=rapport.date_fin.month)
        elif rapport.date_debut:
            filters &= Q(annee__gte=rapport.date_debut.year)
        elif rapport.date_fin:
            filters &= Q(annee__lte=rapport.date_fin.year)

        if rapport.mandat:
            filters &= Q(employe__mandat=rapport.mandat)

        fiches = FicheSalaire.objects.filter(filters).select_related('employe')

        employes = []
        for fiche in fiches:
            employes.append({
                'nom': fiche.employe.nom_complet if hasattr(fiche.employe, 'nom_complet') else str(fiche.employe),
                'brut': fiche.salaire_brut_total or Decimal('0'),
                'charges': fiche.total_cotisations_employe or Decimal('0'),
                'net': fiche.salaire_net or Decimal('0'),
            })

        return {
            'employes': employes,
            'total_brut': sum(e['brut'] for e in employes),
            'total_charges': sum(e['charges'] for e in employes),
            'total_net': sum(e['net'] for e in employes),
        }

    except Exception as e:
        logger.warning(f"Erreur récupération données salaires: {e}")
        return {'employes': [], 'total_brut': 0, 'total_charges': 0, 'total_net': 0}


def _get_donnees_ca_mensuel(rapport):
    """Récupère les données CA mensuel pour le rapport."""
    try:
        Facture = apps.get_model('facturation', 'Facture')

        from django.db.models import Sum
        from django.db.models.functions import TruncMonth

        filters = {
            'date_emission__gte': rapport.date_debut,
            'date_emission__lte': rapport.date_fin,
            'statut__in': ['EMISE', 'ENVOYEE', 'PAYEE', 'PARTIELLEMENT_PAYEE'],
        }
        if rapport.mandat:
            filters['mandat'] = rapport.mandat

        factures = Facture.objects.filter(**filters).annotate(
            mois=TruncMonth('date_emission')
        ).values('mois').annotate(
            ca=Sum('montant_ht')
        ).order_by('mois')

        mois_data = []
        prev_ca = None

        mois_noms = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun', 'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']

        for f in factures:
            ca = f['ca'] or Decimal('0')
            evol = 0
            if prev_ca and prev_ca > 0:
                evol = float((ca - prev_ca) / prev_ca * 100)

            mois_data.append({
                'libelle': mois_noms[f['mois'].month - 1] + ' ' + str(f['mois'].year),
                'ca': ca,
                'evolution': round(evol, 1)
            })
            prev_ca = ca

        return {'mois': mois_data}

    except Exception as e:
        logger.warning(f"Erreur récupération données CA: {e}")
        return {'mois': []}


def _get_donnees_rentabilite(rapport):
    """Récupère les données de rentabilité pour le rapport."""
    try:
        Facture = apps.get_model('facturation', 'Facture')
        EcritureComptable = apps.get_model('comptabilite', 'EcritureComptable')

        from django.db.models import Sum

        # Filtres pour les factures
        facture_filters = {
            'date_emission__gte': rapport.date_debut,
            'date_emission__lte': rapport.date_fin,
            'statut__in': ['EMISE', 'ENVOYEE', 'PAYEE', 'PARTIELLEMENT_PAYEE'],
        }
        if rapport.mandat:
            facture_filters['mandat'] = rapport.mandat

        # Calculer le CA
        ca_data = Facture.objects.filter(**facture_filters).aggregate(
            ca_total=Sum('montant_ht')
        )
        ca_total = ca_data['ca_total'] or Decimal('0')

        # Filtres pour les écritures comptables (charges)
        ecriture_filters = {
            'date_ecriture__gte': rapport.date_debut,
            'date_ecriture__lte': rapport.date_fin,
        }
        if rapport.mandat:
            ecriture_filters['mandat'] = rapport.mandat

        # Récupérer les charges (classes 4, 5, 6, 8 selon plan comptable suisse PME)
        charges_data = EcritureComptable.objects.filter(
            **ecriture_filters,
            compte__classe__in=[4, 5, 6, 8]
        ).aggregate(
            total_debit=Sum('montant_debit'),
            total_credit=Sum('montant_credit')
        )
        total_charges = (charges_data['total_debit'] or Decimal('0')) - (charges_data['total_credit'] or Decimal('0'))

        # Calculs de rentabilité
        marge_brute_montant = ca_total - total_charges if ca_total > 0 else Decimal('0')
        marge_brute_pct = float((marge_brute_montant / ca_total) * 100) if ca_total > 0 else 0

        # Récupérer les produits (classes 3, 7 selon plan comptable suisse PME)
        produits_data = EcritureComptable.objects.filter(
            **ecriture_filters,
            compte__classe__in=[3, 7]
        ).aggregate(
            total_debit=Sum('montant_debit'),
            total_credit=Sum('montant_credit')
        )
        total_produits = (produits_data['total_credit'] or Decimal('0')) - (produits_data['total_debit'] or Decimal('0'))

        # Résultat net et taux de rentabilité
        resultat_net = total_produits - total_charges
        taux_rentabilite = float((resultat_net / total_produits) * 100) if total_produits > 0 else 0

        # Taux de facturation (estimation basée sur CA vs capacité théorique)
        # Si pas de données de temps, on estime basé sur le CA moyen mensuel
        nb_mois = max(1, (rapport.date_fin - rapport.date_debut).days // 30)
        ca_mensuel_moyen = float(ca_total / nb_mois) if nb_mois > 0 else 0

        # Estimation du taux de facturation (basé sur un objectif de 100k/mois par exemple)
        objectif_mensuel = 100000  # À paramétrer selon l'entreprise
        taux_facturation = min(100, (ca_mensuel_moyen / objectif_mensuel) * 100) if objectif_mensuel > 0 else 0

        # Heures facturées (estimation si pas de données de temps)
        heures_estimees = int(float(ca_total) / 150) if ca_total > 0 else 0  # Estimation à 150 CHF/h

        return {
            'ca_total': float(ca_total),
            'charges_total': float(total_charges),
            'marge_brute_montant': float(marge_brute_montant),
            'marge_brute': round(marge_brute_pct, 1),
            'evol_marge': '+0%',  # TODO: Calculer vs période précédente
            'resultat_net': float(resultat_net),
            'taux_rentabilite': round(taux_rentabilite, 1),
            'evol_rentabilite': '+0%',  # TODO: Calculer vs période précédente
            'heures_facturees': heures_estimees,
            'taux_facturation': round(taux_facturation, 1),
        }

    except Exception as e:
        logger.warning(f"Erreur récupération données rentabilité: {e}")
        return {
            'ca_total': 0,
            'charges_total': 0,
            'marge_brute_montant': 0,
            'marge_brute': 0,
            'evol_marge': '+0%',
            'resultat_net': 0,
            'taux_rentabilite': 0,
            'evol_rentabilite': '+0%',
            'heures_facturees': 0,
            'taux_facturation': 0,
        }


def _generer_excel(rapport):
    """Génère le contenu Excel du rapport."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = rapport.type_rapport

    # Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

    # En-tête
    ws['A1'] = rapport.nom
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f"Période: {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}"
    if rapport.mandat:
        ws['A3'] = f"Client: {rapport.mandat.client.raison_sociale}"

    # Données selon le type
    row = 5
    data = _get_donnees_comptables(rapport, rapport.type_rapport.lower())

    # En-têtes colonnes
    headers = ['N° Compte', 'Libellé', 'Débit', 'Crédit', 'Solde']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Données
    row += 1
    for compte in data.get('comptes', []):
        ws.cell(row=row, column=1, value=compte['numero'])
        ws.cell(row=row, column=2, value=compte['libelle'])
        ws.cell(row=row, column=3, value=float(compte.get('debit', 0)))
        ws.cell(row=row, column=4, value=float(compte.get('credit', 0)))
        ws.cell(row=row, column=5, value=float(compte.get('solde', 0)))
        row += 1

    # Auto-fit colonnes
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"{rapport.nom.replace(' ', '_')}_{rapport.date_fin.strftime('%Y%m%d')}.xlsx"
    return buffer.getvalue(), filename


def _generer_csv(rapport):
    """Génère le contenu CSV du rapport."""
    import csv

    output = io.StringIO()
    writer = csv.writer(output, delimiter=';')

    # En-tête
    writer.writerow([rapport.nom])
    writer.writerow([f"Période: {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}"])
    writer.writerow([])

    # Données
    data = _get_donnees_comptables(rapport, rapport.type_rapport.lower())

    writer.writerow(['N° Compte', 'Libellé', 'Débit', 'Crédit', 'Solde'])
    for compte in data.get('comptes', []):
        writer.writerow([
            compte['numero'],
            compte['libelle'],
            float(compte.get('debit', 0)),
            float(compte.get('credit', 0)),
            float(compte.get('solde', 0))
        ])

    content = '\ufeff' + output.getvalue()  # BOM UTF-8
    filename = f"{rapport.nom.replace(' ', '_')}_{rapport.date_fin.strftime('%Y%m%d')}.csv"
    return content.encode('utf-8'), filename


def _generer_html(rapport):
    """Génère le contenu HTML du rapport."""
    data = _get_donnees_comptables(rapport, rapport.type_rapport.lower())

    html = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>{rapport.nom}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; }}
        h1 {{ color: #366092; }}
        .info {{ color: #666; margin-bottom: 20px; }}
        table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
        th {{ background: #366092; color: white; padding: 10px; text-align: left; }}
        td {{ border: 1px solid #ddd; padding: 8px; }}
        tr:nth-child(even) {{ background: #f5f5f5; }}
        .montant {{ text-align: right; }}
    </style>
</head>
<body>
    <h1>{rapport.nom}</h1>
    <div class="info">
        <p>Période: {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}</p>
        {'<p>Client: ' + rapport.mandat.client.raison_sociale + '</p>' if rapport.mandat else ''}
    </div>

    <table>
        <tr>
            <th>N° Compte</th>
            <th>Libellé</th>
            <th>Débit</th>
            <th>Crédit</th>
            <th>Solde</th>
        </tr>
"""

    for compte in data.get('comptes', []):
        html += f"""        <tr>
            <td>{compte['numero']}</td>
            <td>{compte['libelle']}</td>
            <td class="montant">{_format_montant(compte.get('debit', 0))}</td>
            <td class="montant">{_format_montant(compte.get('credit', 0))}</td>
            <td class="montant">{_format_montant(compte.get('solde', 0))}</td>
        </tr>
"""

    html += """    </table>
</body>
</html>"""

    filename = f"{rapport.nom.replace(' ', '_')}_{rapport.date_fin.strftime('%Y%m%d')}.html"
    return html.encode('utf-8'), filename


def _envoyer_rapport_email(rapport):
    """Envoie le rapport par email aux destinataires."""
    from mailing.models import ConfigurationEmail
    from mailing.services import EmailService

    # Vérifier qu'une configuration email existe
    config = ConfigurationEmail.get_default('NOREPLY')
    if not config:
        logger.warning(f"Pas de configuration email pour envoyer le rapport {rapport.id}")
        return

    service = EmailService(configuration=config)

    # Préparer le contenu
    sujet = f"Rapport: {rapport.nom}"
    corps_html = f"""
    <h2>{rapport.nom}</h2>
    <p>Veuillez trouver ci-joint le rapport demandé.</p>
    <p><strong>Période:</strong> {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}</p>
    <p><strong>Type:</strong> {rapport.get_type_rapport_display()}</p>
    <p><strong>Généré le:</strong> {timezone.now().strftime('%d.%m.%Y à %H:%M')}</p>
    <br>
    <p>Cordialement,<br>AltiusOne</p>
    """

    # Pièce jointe
    pieces_jointes = []
    if rapport.fichier:
        pieces_jointes.append({
            'nom': rapport.fichier.name.split('/')[-1],
            'path': rapport.fichier.path
        })

    # Envoyer à chaque destinataire
    for destinataire in rapport.destinataires:
        try:
            service.send_email(
                destinataire=destinataire,
                sujet=sujet,
                corps_html=corps_html,
                pieces_jointes=pieces_jointes,
                utilisateur=rapport.genere_par,
                mandat=rapport.mandat,
                content_type='RAPPORT',
                object_id=str(rapport.id),
                async_send=True
            )
            logger.info(f"Email envoyé à {destinataire} pour le rapport {rapport.id}")
        except Exception as e:
            logger.error(f"Erreur envoi email à {destinataire}: {e}")


@shared_task(bind=True, max_retries=3, default_retry_delay=60)
def envoyer_rapport_email_async(self, rapport_id: str, destinataires: list,
                                  message_personnalise: str = '', expediteur_id: str = None):
    """
    Envoie un rapport par email de manière asynchrone.

    Args:
        rapport_id: UUID du rapport à envoyer
        destinataires: Liste des adresses email
        message_personnalise: Message personnalisé à inclure
        expediteur_id: UUID de l'utilisateur qui envoie
    """
    from analytics.models import Rapport
    from mailing.models import ConfigurationEmail, EmailEnvoye
    from django.contrib.auth import get_user_model

    User = get_user_model()

    try:
        rapport = Rapport.objects.select_related('mandat', 'genere_par').get(id=rapport_id)
    except Rapport.DoesNotExist:
        logger.error(f"Rapport {rapport_id} non trouvé pour envoi email")
        return {'success': False, 'error': 'Rapport non trouvé'}

    # Récupérer l'expéditeur
    expediteur = None
    if expediteur_id:
        try:
            expediteur = User.objects.get(id=expediteur_id)
        except User.DoesNotExist:
            expediteur = rapport.genere_par

    # Vérifier la configuration email
    try:
        config = ConfigurationEmail.objects.filter(
            type_config='SMTP',
            actif=True
        ).first()

        if not config:
            logger.error("Aucune configuration email SMTP active")
            return {'success': False, 'error': 'Pas de configuration email'}

    except Exception as e:
        logger.error(f"Erreur configuration email: {e}")
        return {'success': False, 'error': str(e)}

    # Préparer le contenu de l'email
    sujet = f"Rapport: {rapport.nom}"

    # Corps HTML
    corps_html = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: #366092; color: white; padding: 20px; text-align: center; }}
        .content {{ padding: 20px; background: #f9f9f9; }}
        .info {{ margin: 15px 0; }}
        .info strong {{ color: #366092; }}
        .message {{ background: white; padding: 15px; border-left: 3px solid #366092; margin: 20px 0; }}
        .footer {{ text-align: center; color: #888; font-size: 12px; padding: 20px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>{rapport.nom}</h1>
        </div>
        <div class="content">
            <p>Bonjour,</p>
            <p>Veuillez trouver ci-joint le rapport demandé.</p>

            <div class="info">
                <p><strong>Type de rapport:</strong> {rapport.get_type_rapport_display()}</p>
                <p><strong>Période:</strong> {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}</p>
                <p><strong>Format:</strong> {rapport.format_fichier}</p>
                <p><strong>Généré le:</strong> {rapport.date_generation.strftime('%d.%m.%Y à %H:%M') if rapport.date_generation else 'N/A'}</p>
                {'<p><strong>Client:</strong> ' + rapport.mandat.client.raison_sociale + '</p>' if rapport.mandat else ''}
            </div>

            {'<div class="message"><p><strong>Message:</strong></p><p>' + message_personnalise + '</p></div>' if message_personnalise else ''}

            <p>Cordialement,<br>
            {expediteur.get_full_name() if expediteur and expediteur.get_full_name() else 'AltiusOne'}</p>
        </div>
        <div class="footer">
            <p>Ce message a été envoyé automatiquement par AltiusOne.</p>
        </div>
    </div>
</body>
</html>
"""

    # Corps texte (version plain text)
    corps_texte = f"""
Rapport: {rapport.nom}

Type: {rapport.get_type_rapport_display()}
Période: {rapport.date_debut.strftime('%d.%m.%Y')} - {rapport.date_fin.strftime('%d.%m.%Y')}
Format: {rapport.format_fichier}
{'Client: ' + rapport.mandat.client.raison_sociale if rapport.mandat else ''}

{('Message: ' + message_personnalise) if message_personnalise else ''}

Veuillez trouver le rapport en pièce jointe.

Cordialement,
{expediteur.get_full_name() if expediteur and expediteur.get_full_name() else 'AltiusOne'}
"""

    # Préparer la pièce jointe
    pieces_jointes = []
    if rapport.fichier:
        try:
            pieces_jointes.append({
                'nom': rapport.fichier.name.split('/')[-1],
                'path': rapport.fichier.path
            })
        except Exception as e:
            logger.warning(f"Impossible d'attacher le fichier: {e}")

    # Envoyer à chaque destinataire
    succes = 0
    erreurs = []

    import smtplib
    from email.mime.text import MIMEText
    from email.mime.multipart import MIMEMultipart
    from email.mime.base import MIMEBase
    from email import encoders

    for destinataire in destinataires:
        try:
            # Créer le message
            msg = MIMEMultipart('alternative')
            msg['Subject'] = sujet
            msg['From'] = f"{config.from_name} <{config.email_address}>" if config.from_name else config.email_address
            msg['To'] = destinataire
            if config.reply_to:
                msg['Reply-To'] = config.reply_to

            # Ajouter le corps
            msg.attach(MIMEText(corps_texte, 'plain', 'utf-8'))
            msg.attach(MIMEText(corps_html, 'html', 'utf-8'))

            # Ajouter les pièces jointes
            for pj in pieces_jointes:
                try:
                    with open(pj['path'], 'rb') as f:
                        part = MIMEBase('application', 'octet-stream')
                        part.set_payload(f.read())
                        encoders.encode_base64(part)
                        part.add_header('Content-Disposition', f'attachment; filename="{pj["nom"]}"')
                        msg.attach(part)
                except Exception as e:
                    logger.warning(f"Erreur pièce jointe {pj['nom']}: {e}")

            # Envoyer
            if config.smtp_use_ssl:
                server = smtplib.SMTP_SSL(config.smtp_host, config.smtp_port)
            else:
                server = smtplib.SMTP(config.smtp_host, config.smtp_port)
                if config.smtp_use_tls:
                    server.starttls()

            server.login(config.username, config.password)
            server.sendmail(config.email_address, destinataire, msg.as_string())
            server.quit()

            # Enregistrer l'envoi
            EmailEnvoye.objects.create(
                configuration=config,
                destinataire=destinataire,
                sujet=sujet,
                corps_html=corps_html,
                corps_texte=corps_texte,
                pieces_jointes=pieces_jointes,
                statut='ENVOYE',
                date_envoi=timezone.now(),
                utilisateur=expediteur,
                mandat=rapport.mandat,
                content_type='RAPPORT',
                object_id=str(rapport.id)
            )

            succes += 1
            logger.info(f"Rapport {rapport.id} envoyé à {destinataire}")

        except Exception as e:
            logger.error(f"Erreur envoi à {destinataire}: {e}")
            erreurs.append({'email': destinataire, 'erreur': str(e)})

            # Enregistrer l'échec
            EmailEnvoye.objects.create(
                configuration=config,
                destinataire=destinataire,
                sujet=sujet,
                corps_html=corps_html,
                corps_texte=corps_texte,
                statut='ECHEC',
                erreur=str(e),
                utilisateur=expediteur,
                mandat=rapport.mandat,
                content_type='RAPPORT',
                object_id=str(rapport.id)
            )

    return {
        'success': succes > 0,
        'envoyes': succes,
        'total': len(destinataires),
        'erreurs': erreurs
    }
